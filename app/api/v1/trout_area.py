"""Trout Area (TA) competition management endpoints.

This module provides endpoints for managing Trout Area fishing competitions,
which use head-to-head match-based scoring with qualifier rounds and
knockout brackets.

Key features:
- Event settings configuration
- Lineup generation with multiple pairing algorithms
- Match management and scoring
- Game card submission and validation (self-validation between competitors)
- Knockout bracket generation
- Real-time ranking updates
"""

import re
from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional


def sanitize_filename(name: str) -> str:
    """Sanitize event name for use in filename."""
    # Replace spaces with underscores
    name = name.replace(" ", "_")
    # Remove special characters except underscores and hyphens
    name = re.sub(r"[^a-zA-Z0-9_\-]", "", name)
    # Truncate to reasonable length
    return name[:50]

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from sqlalchemy import func, select, and_, delete, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.dependencies import get_current_user
from app.core.permissions import OrganizerOrAdmin, EventOwnerOrAdmin
from app.core.i18n import get_error_message
from app.core.exceptions import NotFoundError, ValidationError, ConflictError
from app.services.redis_cache import redis_cache
from app.utils.lifecycle_guards import require_modifiable_status, require_draft_status

from app.models.user import UserAccount, UserProfile
from app.models.event import Event, EventStatus
from app.models.enrollment import EventEnrollment, EnrollmentStatus
from app.models.trout_area import (
    TAPointsRule,
    TAEventPointConfig,
    TAEventSettings,
    TALineup,
    TAMatch,
    TAGameCard,
    TAKnockoutBracket,
    TAKnockoutMatch,
    TAQualifierStanding,
    TAMatchOutcome,
    TATournamentPhase,
    TAMatchStatus,
    TAGameCardStatus,
)

from app.schemas.trout_area import (
    # Points Rules
    TAPointsRuleResponse,
    TAPointsRuleUpdate,
    TAGlobalPointDefaultsResponse,
    TAGlobalPointDefaultsUpdate,
    # Event Point Config
    TAEventPointConfigResponse,
    TAEventPointConfigUpdate,
    # Settings
    TAEventSettingsCreate,
    TAEventSettingsUpdate,
    TAEventSettingsResponse,
    # Lineup
    TALineupResponse,
    TALineupListResponse,
    TAGenerateLineupRequest,
    TAGenerateLineupResponse,
    TAGenerateBracketResponse,
    # Match
    TAMatchResponse,
    TAMatchDetailResponse,
    TAMatchListResponse,
    TAMatchResultUpdate,
    # Game Card
    TAGameCardResponse,
    TAGameCardSubmitRequest,
    TAGameCardValidateRequest,
    TAGameCardAdminUpdateRequest,
    TAMyGameCardsResponse,
    # Bracket
    TAKnockoutBracketResponse,
    TABracketGenerateRequest,
    # Standings
    TAQualifierStandingResponse,
    TAQualifierStandingListResponse,
    # Rankings
    TARankingMovementResponse,
    TARankingUpdateResponse,
    # Duration
    TADurationEstimateRequest,
    TADurationEstimateResponse,
    # Algorithm Preview
    TAAlgorithmOption,
    TAAlgorithmPreviewResponse,
    # Schedule
    TARoundResponse,
    TAScheduleResponse,
    # Enums
    PairingAlgorithmAPI,
    TATournamentPhaseAPI,
    TAMatchOutcomeAPI,
    TAMatchStatusAPI,
    TAGameCardStatusAPI,
)

from app.schemas.common import MessageResponse
from app.services.ta_pairing import TAPairingService, PairingAlgorithm
from app.models.club import ClubMembership, MembershipStatus
from app.api.v1.live import live_scoring_service
from app.services.statistics_service import statistics_service

router = APIRouter()


# =============================================================================
# Helper Functions
# =============================================================================


# =============================================================================
# SSE Broadcast Functions for Live Public Leaderboard
# =============================================================================

async def broadcast_ta_leg_complete(
    event_id: int,
    leg_number: int,
    phase: str,
    standings_updated: bool = True,
) -> None:
    """Broadcast leg completion to live scoring subscribers."""
    await live_scoring_service.broadcast(event_id, {
        "type": "ta_leg_complete",
        "event_id": event_id,
        "leg_number": leg_number,
        "phase": phase,
        "standings_updated": standings_updated,
    })


async def broadcast_ta_standings_update(
    event_id: int,
    phase: str,
    top_changes: list[dict] | None = None,
) -> None:
    """Broadcast standings update after recalculation."""
    await live_scoring_service.broadcast(event_id, {
        "type": "ta_standings_update",
        "event_id": event_id,
        "phase": phase,
        "top_changes": top_changes or [],
    })


async def broadcast_ta_phase_advanced(
    event_id: int,
    from_phase: str,
    to_phase: str,
) -> None:
    """Broadcast phase advancement (e.g., qualifier → semifinals)."""
    await live_scoring_service.broadcast(event_id, {
        "type": "ta_phase_advanced",
        "event_id": event_id,
        "from_phase": from_phase,
        "to_phase": to_phase,
    })


async def broadcast_ta_bracket_generated(
    event_id: int,
    semifinalists: list[int],
    requalification_participants: list[int] | None = None,
) -> None:
    """Broadcast knockout bracket generation."""
    await live_scoring_service.broadcast(event_id, {
        "type": "ta_bracket_generated",
        "event_id": event_id,
        "semifinalists": semifinalists,
        "requalification_participants": requalification_participants or [],
    })


async def broadcast_ta_match_result(
    event_id: int,
    match_id: int,
    phase: str,
    leg_number: int,
    competitor_a_id: int,
    competitor_b_id: int,
    competitor_a_catches: int,
    competitor_b_catches: int,
    winner_id: int | None,
) -> None:
    """Broadcast individual match result completion."""
    await live_scoring_service.broadcast(event_id, {
        "type": "ta_match_result",
        "event_id": event_id,
        "match_id": match_id,
        "phase": phase,
        "leg_number": leg_number,
        "competitor_a_id": competitor_a_id,
        "competitor_b_id": competitor_b_id,
        "competitor_a_catches": competitor_a_catches,
        "competitor_b_catches": competitor_b_catches,
        "winner_id": winner_id,
    })


async def _cascade_knockout_update(db: AsyncSession, event_id: int, match: "TAMatch") -> dict | None:
    """
    Cascade updates to downstream phases when a knockout match result is edited.

    - If a requalification match is edited → update semifinal matches with new winners
    - If a semifinal match is edited → update finals matches with new winners/losers

    Returns cascade info or None if no cascade needed.
    """
    from app.models.trout_area import TAMatch, TAGameCard, TAGameCardStatus

    phase = match.phase

    # Only cascade for requalification and semifinal phases
    if phase == TATournamentPhase.REQUALIFICATION.value:
        # Get all requalification matches to check if all are completed
        requalification_query = select(TAMatch).where(
            TAMatch.event_id == event_id,
            TAMatch.phase == TATournamentPhase.REQUALIFICATION.value,
        ).order_by(TAMatch.match_number)
        result = await db.execute(requalification_query)
        requalification_matches = result.scalars().all()

        # Only cascade if ALL requalification matches are completed
        all_completed = all(m.status == TAMatchStatus.COMPLETED.value for m in requalification_matches)
        if not all_completed:
            return None

        # Determine winners from all requalification matches
        requalification_winners = []
        for m in requalification_matches:
            a_catches = m.competitor_a_catches or 0
            b_catches = m.competitor_b_catches or 0
            if a_catches > b_catches:
                winner_id = m.competitor_a_id
            elif b_catches > a_catches:
                winner_id = m.competitor_b_id
            else:
                winner_id = m.competitor_a_id  # Tie goes to higher seed
            requalification_winners.append(winner_id)

        # Get semifinal matches
        sf_query = select(TAMatch).where(
            TAMatch.event_id == event_id,
            TAMatch.phase == TATournamentPhase.SEMIFINAL.value,
        ).order_by(TAMatch.match_number)
        sf_result = await db.execute(sf_query)
        semifinals = list(sf_result.scalars().all())

        if len(semifinals) < 2 or len(requalification_winners) < 2:
            return None

        # Update semifinal competitor_b with requalification winners
        # SF1: Seed 1 vs Winner of 3v5 (requalification match #1)
        # SF2: Seed 2 vs Winner of 4v6 (requalification match #2)
        old_sf1_b = semifinals[0].competitor_b_id
        old_sf2_b = semifinals[1].competitor_b_id
        new_sf1_b = requalification_winners[0]  # Winner of match 1 (3v5)
        new_sf2_b = requalification_winners[1]  # Winner of match 2 (4v6)

        # Helper to ensure game cards exist for both SF competitors
        async def ensure_sf_game_cards(sf_match: TAMatch, comp_a_id: int, comp_b_id: int):
            """Ensure game cards exist for both semifinal competitors."""
            for user_id, opponent_id in [(comp_a_id, comp_b_id), (comp_b_id, comp_a_id)]:
                if user_id:
                    existing = await db.execute(select(TAGameCard).where(
                        TAGameCard.match_id == sf_match.id,
                        TAGameCard.user_id == user_id,
                    ))
                    existing_card = existing.scalar_one_or_none()
                    if existing_card:
                        # Update opponent if changed
                        if existing_card.opponent_id != opponent_id:
                            existing_card.opponent_id = opponent_id
                    else:
                        # Create new card
                        db.add(TAGameCard(
                            event_id=event_id,
                            match_id=sf_match.id,
                            leg_number=sf_match.leg_number,
                            user_id=user_id,
                            opponent_id=opponent_id,
                            my_seat=1,
                            opponent_seat=1,
                            is_ghost_opponent=False,
                            status=TAGameCardStatus.DRAFT.value,
                        ))

        if old_sf1_b != new_sf1_b:
            semifinals[0].competitor_b_id = new_sf1_b
            # Delete old placeholder's game card
            if old_sf1_b:
                await db.execute(delete(TAGameCard).where(
                    TAGameCard.match_id == semifinals[0].id,
                    TAGameCard.user_id == old_sf1_b,
                ))
            # Ensure both competitors have game cards
            await ensure_sf_game_cards(
                semifinals[0],
                semifinals[0].competitor_a_id,
                new_sf1_b
            )

        if old_sf2_b != new_sf2_b:
            semifinals[1].competitor_b_id = new_sf2_b
            # Delete old placeholder's game card
            if old_sf2_b:
                await db.execute(delete(TAGameCard).where(
                    TAGameCard.match_id == semifinals[1].id,
                    TAGameCard.user_id == old_sf2_b,
                ))
            # Ensure both competitors have game cards
            await ensure_sf_game_cards(
                semifinals[1],
                semifinals[1].competitor_a_id,
                new_sf2_b
            )

        return {"cascaded_to": "semifinals", "updated_winners": requalification_winners}

    elif phase == TATournamentPhase.SEMIFINAL.value:
        # Get all semifinal matches to check if all are completed
        sf_query = select(TAMatch).where(
            TAMatch.event_id == event_id,
            TAMatch.phase == TATournamentPhase.SEMIFINAL.value,
        ).order_by(TAMatch.match_number)
        sf_result = await db.execute(sf_query)
        semifinals = list(sf_result.scalars().all())

        all_completed = all(m.status == TAMatchStatus.COMPLETED.value for m in semifinals)
        if not all_completed or len(semifinals) < 2:
            return None

        # Determine winners and losers from semifinals
        sf_results = []
        for sf in semifinals:
            a_catches = sf.competitor_a_catches or 0
            b_catches = sf.competitor_b_catches or 0
            if a_catches > b_catches:
                winner_id = sf.competitor_a_id
                loser_id = sf.competitor_b_id
            elif b_catches > a_catches:
                winner_id = sf.competitor_b_id
                loser_id = sf.competitor_a_id
            else:
                winner_id = sf.competitor_a_id
                loser_id = sf.competitor_b_id
            sf_results.append({"winner": winner_id, "loser": loser_id})

        # Get finals matches (Grand Final and Small Final)
        finals_query = select(TAMatch).where(
            TAMatch.event_id == event_id,
            TAMatch.phase.in_([
                TATournamentPhase.FINAL_GRAND.value,
                TATournamentPhase.FINAL_SMALL.value,
            ]),
        )
        finals_result = await db.execute(finals_query)
        finals = {m.phase: m for m in finals_result.scalars().all()}

        # Update Grand Final with winners
        grand_final = finals.get(TATournamentPhase.FINAL_GRAND.value)
        if grand_final:
            old_gf_a = grand_final.competitor_a_id
            old_gf_b = grand_final.competitor_b_id
            new_gf_a = sf_results[0]["winner"]
            new_gf_b = sf_results[1]["winner"]
            grand_final.competitor_a_id = new_gf_a
            grand_final.competitor_b_id = new_gf_b

            # Delete old game cards and create new ones for grand final
            if old_gf_a and old_gf_a != new_gf_a:
                await db.execute(delete(TAGameCard).where(
                    TAGameCard.match_id == grand_final.id,
                    TAGameCard.user_id == old_gf_a,
                ))
            if old_gf_b and old_gf_b != new_gf_b:
                await db.execute(delete(TAGameCard).where(
                    TAGameCard.match_id == grand_final.id,
                    TAGameCard.user_id == old_gf_b,
                ))

            # Create/update game cards for grand final competitors
            for user_id, opponent_id in [(new_gf_a, new_gf_b), (new_gf_b, new_gf_a)]:
                if user_id:
                    existing = await db.execute(select(TAGameCard).where(
                        TAGameCard.match_id == grand_final.id,
                        TAGameCard.user_id == user_id,
                    ))
                    existing_card = existing.scalar_one_or_none()
                    if existing_card:
                        existing_card.opponent_id = opponent_id
                    else:
                        db.add(TAGameCard(
                            event_id=event_id,
                            match_id=grand_final.id,
                            leg_number=grand_final.leg_number,
                            user_id=user_id,
                            opponent_id=opponent_id,
                            my_seat=1,
                            opponent_seat=1,
                            is_ghost_opponent=False,
                            status=TAGameCardStatus.DRAFT.value,
                        ))

        # Update Small Final with losers
        small_final = finals.get(TATournamentPhase.FINAL_SMALL.value)
        if small_final:
            old_sf_a = small_final.competitor_a_id
            old_sf_b = small_final.competitor_b_id
            new_sf_a = sf_results[0]["loser"]
            new_sf_b = sf_results[1]["loser"]
            small_final.competitor_a_id = new_sf_a
            small_final.competitor_b_id = new_sf_b

            # Delete old game cards and create new ones for small final
            if old_sf_a and old_sf_a != new_sf_a:
                await db.execute(delete(TAGameCard).where(
                    TAGameCard.match_id == small_final.id,
                    TAGameCard.user_id == old_sf_a,
                ))
            if old_sf_b and old_sf_b != new_sf_b:
                await db.execute(delete(TAGameCard).where(
                    TAGameCard.match_id == small_final.id,
                    TAGameCard.user_id == old_sf_b,
                ))

            # Create/update game cards for small final competitors
            for user_id, opponent_id in [(new_sf_a, new_sf_b), (new_sf_b, new_sf_a)]:
                if user_id:
                    existing = await db.execute(select(TAGameCard).where(
                        TAGameCard.match_id == small_final.id,
                        TAGameCard.user_id == user_id,
                    ))
                    existing_card = existing.scalar_one_or_none()
                    if existing_card:
                        existing_card.opponent_id = opponent_id
                    else:
                        db.add(TAGameCard(
                            event_id=event_id,
                            match_id=small_final.id,
                            leg_number=small_final.leg_number,
                            user_id=user_id,
                            opponent_id=opponent_id,
                            my_seat=1,
                            opponent_seat=1,
                            is_ghost_opponent=False,
                            status=TAGameCardStatus.DRAFT.value,
                        ))

        return {
            "cascaded_to": "finals",
            "grand_final_a": sf_results[0]["winner"],
            "grand_final_b": sf_results[1]["winner"],
            "small_final_a": sf_results[0]["loser"],
            "small_final_b": sf_results[1]["loser"],
        }

    # Check if both finals are completed → update final_standings
    if phase in [TATournamentPhase.FINAL_GRAND.value, TATournamentPhase.FINAL_SMALL.value]:
        finals_query = select(TAMatch).where(
            TAMatch.event_id == event_id,
            TAMatch.phase.in_([
                TATournamentPhase.FINAL_GRAND.value,
                TATournamentPhase.FINAL_SMALL.value,
            ]),
        )
        finals_result = await db.execute(finals_query)
        finals = list(finals_result.scalars().all())

        all_completed = len(finals) == 2 and all(m.status == TAMatchStatus.COMPLETED.value for m in finals)
        if all_completed:
            # Both finals done - update knockout bracket final_standings
            grand_final = next((m for m in finals if m.phase == TATournamentPhase.FINAL_GRAND.value), None)
            small_final = next((m for m in finals if m.phase == TATournamentPhase.FINAL_SMALL.value), None)

            final_standings = {}
            if grand_final:
                gf_a = grand_final.competitor_a_catches or 0
                gf_b = grand_final.competitor_b_catches or 0
                if gf_a > gf_b:
                    final_standings["1"] = grand_final.competitor_a_id
                    final_standings["2"] = grand_final.competitor_b_id
                elif gf_b > gf_a:
                    final_standings["1"] = grand_final.competitor_b_id
                    final_standings["2"] = grand_final.competitor_a_id
                else:  # Tie - competitor_a (higher seed) wins
                    final_standings["1"] = grand_final.competitor_a_id
                    final_standings["2"] = grand_final.competitor_b_id

            if small_final:
                sf_a = small_final.competitor_a_catches or 0
                sf_b = small_final.competitor_b_catches or 0
                if sf_a > sf_b:
                    final_standings["3"] = small_final.competitor_a_id
                    final_standings["4"] = small_final.competitor_b_id
                elif sf_b > sf_a:
                    final_standings["3"] = small_final.competitor_b_id
                    final_standings["4"] = small_final.competitor_a_id
                else:  # Tie - competitor_a (higher seed) wins
                    final_standings["3"] = small_final.competitor_a_id
                    final_standings["4"] = small_final.competitor_b_id

            # Update knockout bracket
            from app.models.trout_area import TAKnockoutBracket
            bracket_query = select(TAKnockoutBracket).where(TAKnockoutBracket.event_id == event_id)
            bracket_result = await db.execute(bracket_query)
            bracket = bracket_result.scalar_one_or_none()
            if bracket:
                bracket.final_standings = final_standings
                bracket.is_completed = True

            return {
                "cascaded_to": "final_standings",
                "final_standings": final_standings,
            }

    return None


async def get_ta_event(
    event_id: int,
    db: AsyncSession,
    request: Request,
    require_settings: bool = True,
) -> Event:
    """Get event and verify it's a TA competition."""
    query = (
        select(Event)
        .options(selectinload(Event.event_type))
        .options(selectinload(Event.ta_settings))
        .where(Event.id == event_id, Event.is_deleted == False)
    )
    result = await db.execute(query)
    event = result.scalar_one_or_none()

    if not event:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=get_error_message("ta_event_not_found", request),
        )

    # Verify event type is TA
    if event.event_type and event.event_type.code not in ["trout_area"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=get_error_message("ta_invalid_event_type", request),
        )

    if require_settings and not event.ta_settings:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=get_error_message("ta_settings_not_found", request),
        )

    return event


async def get_user_active_club_id(db: AsyncSession, user_id: int) -> int | None:
    """Get user's active club_id (first active membership).

    Args:
        db: Database session
        user_id: User ID to get club for

    Returns:
        Club ID if user has active membership, None otherwise
    """
    query = (
        select(ClubMembership.club_id)
        .where(
            ClubMembership.user_id == user_id,
            ClubMembership.status == MembershipStatus.ACTIVE.value,
        )
        .order_by(ClubMembership.joined_at.desc())
        .limit(1)
    )
    result = await db.execute(query)
    return result.scalar_one_or_none()


async def get_club_ids_for_users(db: AsyncSession, user_ids: list[int]) -> dict[int, int | None]:
    """Get active club_ids for multiple users efficiently.

    Args:
        db: Database session
        user_ids: List of user IDs

    Returns:
        Dict mapping user_id -> club_id (or None if no active membership)
    """
    if not user_ids:
        return {}

    # Get all active memberships for these users
    query = (
        select(ClubMembership.user_id, ClubMembership.club_id)
        .where(
            ClubMembership.user_id.in_(user_ids),
            ClubMembership.status == MembershipStatus.ACTIVE.value,
        )
        .order_by(ClubMembership.user_id, ClubMembership.joined_at.desc())
        .distinct(ClubMembership.user_id)
    )
    result = await db.execute(query)
    rows = result.all()

    # Build lookup dict
    club_lookup = {row.user_id: row.club_id for row in rows}

    # Fill in None for users without active membership
    return {user_id: club_lookup.get(user_id) for user_id in user_ids}


def map_pairing_algorithm(api_algo: PairingAlgorithmAPI) -> PairingAlgorithm:
    """Map API enum to service enum."""
    mapping = {
        PairingAlgorithmAPI.ROUND_ROBIN_FULL: PairingAlgorithm.ROUND_ROBIN_FULL,
        PairingAlgorithmAPI.ROUND_ROBIN_HALF: PairingAlgorithm.ROUND_ROBIN_HALF,
        PairingAlgorithmAPI.ROUND_ROBIN_CUSTOM: PairingAlgorithm.ROUND_ROBIN_CUSTOM,
        PairingAlgorithmAPI.SIMPLE_PAIRS: PairingAlgorithm.SIMPLE_PAIRS,
    }
    return mapping[api_algo]


async def update_standings_for_match(
    db: AsyncSession,
    match: TAMatch,
    point_config: Optional[TAEventPointConfig] = None,
    skip_rank_recalculation: bool = False,
) -> dict:
    """
    Update standings for both competitors after a match completes.

    This function:
    1. Gets or creates standings for both players
    2. Updates totals (points, catches, W/T/L breakdown)
    3. Updates leg_results JSONB
    4. Checks if leg is complete - only then recalculates ranks (Story 12.1 optimization)
    5. Broadcasts SSE event when leg completes

    Args:
        db: Database session
        match: The completed match
        point_config: Optional custom point configuration
        skip_rank_recalculation: Skip rank recalculation (for batch operations)

    Returns:
        dict with leg_complete status and leg_number
    """
    if match.status != TAMatchStatus.COMPLETED.value:
        return {"leg_complete": False, "leg_number": None, "ranks_updated": False}

    # Get point values (default if no config)
    point_values = {
        "V": Decimal("3.0"),
        "T": Decimal("1.5"),
        "T0": Decimal("1.0"),
        "L": Decimal("0.5"),
        "L0": Decimal("0.0"),
    }

    if point_config:
        point_values = {
            "V": point_config.victory_points,
            "T": point_config.tie_points,
            "T0": point_config.tie_zero_points,
            "L": point_config.loss_points,
            "L0": point_config.loss_zero_points,
        }

    # Get enrollments for both users
    enrollment_query = select(EventEnrollment).where(
        EventEnrollment.event_id == match.event_id,
        EventEnrollment.user_id.in_([match.competitor_a_id, match.competitor_b_id]),
        EventEnrollment.status == "approved",
    )
    enrollment_result = await db.execute(enrollment_query)
    enrollments = {e.user_id: e for e in enrollment_result.scalars().all()}

    # Process each competitor
    for competitor_id, catches, outcome_code in [
        (match.competitor_a_id, match.competitor_a_catches or 0, match.competitor_a_outcome_code),
        (match.competitor_b_id, match.competitor_b_catches or 0, match.competitor_b_outcome_code),
    ]:
        if not competitor_id or competitor_id not in enrollments:
            continue

        enrollment = enrollments[competitor_id]
        points = point_values.get(outcome_code, Decimal("0"))

        # Get or create standing
        standing_query = select(TAQualifierStanding).where(
            TAQualifierStanding.event_id == match.event_id,
            TAQualifierStanding.user_id == competitor_id,
        )
        standing_result = await db.execute(standing_query)
        standing = standing_result.scalar_one_or_none()

        if not standing:
            standing = TAQualifierStanding(
                event_id=match.event_id,
                user_id=competitor_id,
                enrollment_id=enrollment.id,
                total_points=Decimal("0"),
                total_matches=0,
                total_victories=0,
                total_ties=0,
                total_losses=0,
                total_fish_caught=0,
                ties_with_fish=0,
                ties_without_fish=0,
                losses_with_fish=0,
                losses_without_fish=0,
                leg_results={},
            )
            db.add(standing)

        # Update totals
        standing.total_points += points
        standing.total_matches += 1
        standing.total_fish_caught += catches

        # Update W/T/L counts
        if outcome_code == "V":
            standing.total_victories += 1
        elif outcome_code == "T":
            standing.total_ties += 1
            standing.ties_with_fish += 1
        elif outcome_code == "T0":
            standing.total_ties += 1
            standing.ties_without_fish += 1
        elif outcome_code == "L":
            standing.total_losses += 1
            standing.losses_with_fish += 1
        elif outcome_code == "L0":
            standing.total_losses += 1
            standing.losses_without_fish += 1

        # Update leg_results JSONB
        leg_key = str(match.leg_number)
        leg_results = standing.leg_results or {}

        if leg_key not in leg_results:
            leg_results[leg_key] = {
                "points": 0,
                "victories": 0,
                "ties": 0,
                "losses": 0,
                "fish": 0,
            }

        leg_results[leg_key]["points"] = float(leg_results[leg_key].get("points", 0)) + float(points)
        leg_results[leg_key]["fish"] = leg_results[leg_key].get("fish", 0) + catches

        if outcome_code == "V":
            leg_results[leg_key]["victories"] = leg_results[leg_key].get("victories", 0) + 1
        elif outcome_code in ["T", "T0"]:
            leg_results[leg_key]["ties"] = leg_results[leg_key].get("ties", 0) + 1
        elif outcome_code in ["L", "L0"]:
            leg_results[leg_key]["losses"] = leg_results[leg_key].get("losses", 0) + 1

        standing.leg_results = leg_results
        standing.updated_at = datetime.now(timezone.utc)

    await db.flush()

    # Story 12.1: Only recalculate ranks when leg is complete (optimization)
    leg_complete = False
    ranks_updated = False

    if not skip_rank_recalculation:
        from app.services.ta_ranking import TARankingService
        ranking_service = TARankingService(db)

        # Check if this leg is now complete
        leg_complete = await ranking_service.is_leg_complete(
            match.event_id,
            match.leg_number,
            match.phase,
        )

        if leg_complete:
            # Leg is complete - recalculate ranks once for the whole leg
            await recalculate_event_ranks(db, match.event_id)
            ranks_updated = True

            # Broadcast SSE event for leg completion
            try:
                await redis_cache.publish_sse_event(match.event_id, {
                    "type": "ta_leg_complete",
                    "event_id": match.event_id,
                    "leg_number": match.leg_number,
                    "phase": match.phase,
                    "message": f"Leg {match.leg_number} completed - standings updated",
                })
            except Exception as e:
                logger.error(f"Failed to broadcast ta_leg_complete SSE: {e}")

    return {
        "leg_complete": leg_complete,
        "leg_number": match.leg_number,
        "phase": match.phase,
        "ranks_updated": ranks_updated,
    }


async def recalculate_event_ranks(db: AsyncSession, event_id: int) -> None:
    """
    Recalculate ranks for all participants in an event.

    Tie-break order:
    1) points desc
    2) catches desc
    3) victories desc
    4) ties_with_fish desc
    5) ties_without_fish desc
    6) losses_with_fish asc
    7) losses_without_fish asc
    """
    # Get all standings for the event
    standings_query = select(TAQualifierStanding).where(
        TAQualifierStanding.event_id == event_id
    )
    result = await db.execute(standings_query)
    standings = list(result.scalars().all())

    if not standings:
        return

    # Sort by tie-break rules
    standings.sort(key=lambda s: (
        -float(s.total_points),  # Higher points first
        -s.total_fish_caught,    # Higher catches first
        -s.total_victories,      # Higher victories first
        -s.ties_with_fish,       # Higher ties with fish first
        -s.ties_without_fish,    # Higher ties without fish first
        s.losses_with_fish,      # Lower losses with fish first
        s.losses_without_fish,   # Lower losses without fish first
    ))

    # Assign ranks (handle ties)
    current_rank = 1
    for i, standing in enumerate(standings):
        if i > 0:
            prev = standings[i - 1]
            # Check if tied with previous (same points and catches at minimum)
            if (standing.total_points == prev.total_points and
                standing.total_fish_caught == prev.total_fish_caught and
                standing.total_victories == prev.total_victories):
                # Same rank as previous
                pass
            else:
                current_rank = i + 1

        standing.rank = current_rank

    # Update knockout qualification based on settings
    settings_query = select(TAEventSettings).where(TAEventSettings.event_id == event_id)
    settings_result = await db.execute(settings_query)
    settings = settings_result.scalar_one_or_none()

    knockout_qualifiers = settings.knockout_qualifiers if settings else 4

    for standing in standings:
        standing.qualifies_for_knockout = standing.rank <= knockout_qualifiers

    await db.flush()


# =============================================================================
# Points Rules Endpoints
# =============================================================================

@router.get("/points-rules", response_model=list[TAPointsRuleResponse])
async def list_points_rules(
    db: AsyncSession = Depends(get_db),
) -> list[TAPointsRule]:
    """
    List all TA points rules.

    Returns the standard point values for match outcomes:
    - V (Victory): 3.0 points
    - T (Tie with fish): 1.5 points
    - T0 (Tie no fish): 1.0 points
    - L (Loss with fish): 0.5 points
    - L0 (Loss no fish): 0.0 points
    """
    query = (
        select(TAPointsRule)
        .where(TAPointsRule.is_active == True)
        .order_by(TAPointsRule.points.desc())
    )
    result = await db.execute(query)
    return result.scalars().all()


async def get_global_point_defaults(db: AsyncSession) -> dict:
    """
    Helper function to get global point defaults from TAPointsRule table.
    Returns a dict with victory_points, tie_points, etc.
    Falls back to hardcoded defaults if rules don't exist.
    """
    # Map of rule codes to field names
    code_to_field = {
        "V": "victory_points",
        "T": "tie_points",
        "T0": "tie_zero_points",
        "L": "loss_points",
        "L0": "loss_zero_points",
    }

    # Default values as fallback
    defaults = {
        "victory_points": Decimal("3.0"),
        "tie_points": Decimal("1.5"),
        "tie_zero_points": Decimal("1.0"),
        "loss_points": Decimal("0.5"),
        "loss_zero_points": Decimal("0.0"),
    }

    # Fetch all active rules
    query = select(TAPointsRule).where(TAPointsRule.is_active == True)
    result = await db.execute(query)
    rules = result.scalars().all()

    # Build result from rules
    result_dict = dict(defaults)  # Start with defaults
    for rule in rules:
        if rule.code in code_to_field:
            result_dict[code_to_field[rule.code]] = rule.points

    return result_dict


@router.get("/point-defaults", response_model=TAGlobalPointDefaultsResponse)
async def get_point_defaults(
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get global point defaults for TA competitions.

    These are the default point values used when an event doesn't have
    custom point configuration. Returns values from TAPointsRule table.

    - V (Victory): default 3.0 points
    - T (Tie with fish): default 1.5 points
    - T0 (Tie no fish): default 1.0 points
    - L (Loss with fish): default 0.5 points
    - L0 (Loss no fish): default 0.0 points
    """
    return await get_global_point_defaults(db)


@router.put("/point-defaults", response_model=TAGlobalPointDefaultsResponse)
async def update_point_defaults(
    data: TAGlobalPointDefaultsUpdate,
    request: Request,
    current_user: UserAccount = Depends(OrganizerOrAdmin),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Update global point defaults for TA competitions.

    Only administrators can modify global point defaults.
    These defaults are used for new events when no custom config is set.

    Point values must follow logical ordering:
    victory >= tie >= tie_zero >= loss >= loss_zero
    """
    # Map of field names to rule codes
    field_to_code = {
        "victory_points": "V",
        "tie_points": "T",
        "tie_zero_points": "T0",
        "loss_points": "L",
        "loss_zero_points": "L0",
    }

    # Get update data (only non-None values)
    update_data = data.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        if value is None:
            continue
        code = field_to_code.get(field)
        if not code:
            continue

        # Find and update the rule
        query = select(TAPointsRule).where(TAPointsRule.code == code)
        result = await db.execute(query)
        rule = result.scalar_one_or_none()

        if rule:
            rule.points = value
        # If rule doesn't exist, we could create it, but for now just skip

    await db.commit()

    # Return updated defaults
    return await get_global_point_defaults(db)


# =============================================================================
# Event Point Config Endpoints (Per-event customizable point values)
# =============================================================================

@router.get("/events/{event_id}/point-config", response_model=TAEventPointConfigResponse)
async def get_ta_point_config(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get point configuration for a TA event.

    Returns the event's custom point values if configured,
    or the global default values from TAPointsRule table.
    """
    # Verify event exists and is TA
    await get_ta_event(event_id, db, request, require_settings=False)

    # Get custom config if exists
    query = select(TAEventPointConfig).where(TAEventPointConfig.event_id == event_id)
    result = await db.execute(query)
    config = result.scalar_one_or_none()

    if not config:
        # Return global defaults from TAPointsRule table
        defaults = await get_global_point_defaults(db)
        defaults["is_default"] = True
        return defaults

    return {
        "victory_points": config.victory_points,
        "tie_points": config.tie_points,
        "tie_zero_points": config.tie_zero_points,
        "loss_points": config.loss_points,
        "loss_zero_points": config.loss_zero_points,
        "is_default": False,
    }


@router.put("/events/{event_id}/point-config", response_model=TAEventPointConfigResponse)
async def update_ta_point_config(
    event_id: int,
    data: TAEventPointConfigUpdate,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Update point configuration for a TA event.

    Only the event organizer or administrator can modify point values.
    Creates a new config if one doesn't exist, or updates the existing one.

    Point values must follow logical ordering:
    victory >= tie >= tie_zero >= loss >= loss_zero
    """
    # Verify event exists and is TA
    event = await get_ta_event(event_id, db, request, require_settings=False)

    # Guard: Point config can only be modified in Draft status
    require_draft_status(event, action="modify point configuration")

    # Get or create config
    query = select(TAEventPointConfig).where(TAEventPointConfig.event_id == event_id)
    result = await db.execute(query)
    config = result.scalar_one_or_none()

    if config:
        # Update existing config
        update_data = data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            if value is not None:
                setattr(config, key, value)
    else:
        # Create new config with provided values (or defaults)
        config = TAEventPointConfig(
            event_id=event_id,
            victory_points=data.victory_points if data.victory_points is not None else Decimal("3.0"),
            tie_points=data.tie_points if data.tie_points is not None else Decimal("1.5"),
            tie_zero_points=data.tie_zero_points if data.tie_zero_points is not None else Decimal("1.0"),
            loss_points=data.loss_points if data.loss_points is not None else Decimal("0.5"),
            loss_zero_points=data.loss_zero_points if data.loss_zero_points is not None else Decimal("0.0"),
        )
        db.add(config)

    await db.commit()
    await db.refresh(config)

    return {
        "victory_points": config.victory_points,
        "tie_points": config.tie_points,
        "tie_zero_points": config.tie_zero_points,
        "loss_points": config.loss_points,
        "loss_zero_points": config.loss_zero_points,
        "is_default": False,
    }


@router.delete("/events/{event_id}/point-config", response_model=MessageResponse)
async def reset_ta_point_config(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Reset point configuration to defaults by deleting custom config.

    After reset, the event will use the global default point values.
    """
    # Verify event exists and is TA
    event = await get_ta_event(event_id, db, request, require_settings=False)

    # Guard: Point config can only be modified in Draft status
    require_draft_status(event, action="reset point configuration")

    # Delete custom config if exists
    query = select(TAEventPointConfig).where(TAEventPointConfig.event_id == event_id)
    result = await db.execute(query)
    config = result.scalar_one_or_none()

    if config:
        await db.delete(config)
        await db.commit()
        return {"message": "Point configuration reset to defaults"}

    return {"message": "Event was already using default point values"}


# =============================================================================
# Event Settings Endpoints
# =============================================================================

@router.post(
    "/events/{event_id}/settings",
    response_model=TAEventSettingsResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_event_settings(
    event_id: int,
    data: TAEventSettingsCreate,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> TAEventSettings:
    """
    Create TA settings for an event.

    Only the event owner or administrator can configure TA settings.
    Settings must be created before generating lineups.
    """
    # Get event without requiring existing settings
    event = await get_ta_event(event_id, db, request, require_settings=False)

    # Check if settings already exist
    if event.ta_settings:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="TA settings already exist for this event. Use PATCH to update.",
        )

    settings = TAEventSettings(
        event_id=event_id,
        match_duration_minutes=data.match_duration_minutes,
        number_of_legs=data.legs_per_match,
        max_rounds_per_leg=data.matches_per_leg or 1,
        has_knockout_stage=data.has_knockout_bracket,
        knockout_qualifiers=data.qualification_top_n,
        requalification_slots=data.requalification_spots,
        has_requalification=data.enable_requalification,
        is_team_event=data.enable_team_scoring,
        team_size=data.team_size,
        additional_rules=data.additional_rules or {},
    )

    db.add(settings)

    # Also create default point config for this event
    # This ensures every TA event has its own point configuration from the start
    point_defaults = await get_global_point_defaults(db)
    point_config = TAEventPointConfig(
        event_id=event_id,
        victory_points=point_defaults["victory_points"],
        tie_points=point_defaults["tie_points"],
        tie_zero_points=point_defaults["tie_zero_points"],
        loss_points=point_defaults["loss_points"],
        loss_zero_points=point_defaults["loss_zero_points"],
    )
    db.add(point_config)

    await db.commit()
    await db.refresh(settings)

    return settings


@router.get("/events/{event_id}/settings", response_model=TAEventSettingsResponse)
async def get_event_settings(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> TAEventSettings:
    """Get TA settings for an event."""
    event = await get_ta_event(event_id, db, request)
    return event.ta_settings


@router.patch("/events/{event_id}/settings", response_model=TAEventSettingsResponse)
async def update_event_settings(
    event_id: int,
    data: TAEventSettingsUpdate,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> TAEventSettings:
    """
    Update TA settings for an event.

    Note: Some settings cannot be changed after lineups are generated.
    - has_knockout_bracket can only be changed in DRAFT status
    """
    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    update_data = data.model_dump(exclude_unset=True)

    # Guard: has_knockout_bracket can only be modified in DRAFT status
    if "has_knockout_bracket" in update_data:
        require_draft_status(event, action="modify knockout bracket setting")

    # Field mappings: API field name -> model field name
    field_mappings = {
        "has_knockout_bracket": "has_knockout_stage",
    }

    for field, value in update_data.items():
        # Map API field names to model field names
        model_field = field_mappings.get(field, field)

        if hasattr(settings, model_field):
            if field == "pairing_algorithm" and value:
                setattr(settings, model_field, value.value if hasattr(value, "value") else value)
            else:
                setattr(settings, model_field, value)

    settings.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(settings)

    return settings


# =============================================================================
# Schedule Endpoints (for mobile app)
# =============================================================================

@router.get("/events/{event_id}/schedule", response_model=TAScheduleResponse)
async def get_event_schedule(
    event_id: int,
    phase: Optional[TATournamentPhaseAPI] = None,
    request: Request = None,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get TA event schedule with all legs (manse) and matches.

    This endpoint provides a structured view of the competition schedule,
    organized by legs with match details.
    """
    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    # Build match query
    match_query = (
        select(TAMatch)
        .options(
            selectinload(TAMatch.competitor_a).selectinload(UserAccount.profile),
            selectinload(TAMatch.competitor_b).selectinload(UserAccount.profile),
        )
        .where(TAMatch.event_id == event_id)
        .order_by(TAMatch.round_number, TAMatch.match_number)
    )

    if phase:
        match_query = match_query.where(TAMatch.phase == phase.value)

    result = await db.execute(match_query)
    matches = result.scalars().all()

    if not matches:
        return {
            "legs": [],
            "current_leg": None,
            "total_legs": 0,
            "matches_completed": 0,
            "total_matches": 0,
            # Backwards compatibility
            "rounds": [],
            "current_round": None,
            "total_rounds": 0,
        }

    # Group matches by (leg_number, phase) - different phases can have same leg numbers
    # Key: (leg_num, phase) -> list of matches
    legs_dict: dict[tuple[int, str], list] = {}
    total_completed = 0
    current_leg_key: tuple[int, str] | None = None

    for match in matches:
        leg_num = match.round_number  # round_number in model = leg
        match_phase = match.phase or TATournamentPhase.QUALIFIER.value
        leg_key = (leg_num, match_phase)

        if leg_key not in legs_dict:
            legs_dict[leg_key] = []

        # Build match response with player names
        match_data = TAMatchResponse(
            id=match.id,
            event_id=match.event_id,
            leg_number=match.round_number,  # Map to leg_number for API
            match_number=match.match_number,
            phase=TATournamentPhaseAPI(match.phase) if match.phase else TATournamentPhaseAPI.QUALIFIER,
            player_a_id=match.competitor_a_id,
            player_b_id=match.competitor_b_id,
            seat_a=match.seat_a,
            seat_b=match.seat_b,
            player_a_catches=match.competitor_a_catches or 0,
            player_b_catches=match.competitor_b_catches or 0,
            player_a_points=match.competitor_a_points or 0,
            player_b_points=match.competitor_b_points or 0,
            player_a_outcome=TAMatchOutcomeAPI(match.competitor_a_outcome_code) if match.competitor_a_outcome_code else None,
            player_b_outcome=TAMatchOutcomeAPI(match.competitor_b_outcome_code) if match.competitor_b_outcome_code else None,
            status=TAMatchStatusAPI(match.status) if match.status else TAMatchStatusAPI.PENDING,
            started_at=match.started_at,
            completed_at=match.completed_at,
            created_at=match.created_at,
            player_a_name=match.competitor_a.profile.full_name if match.competitor_a and match.competitor_a.profile else None,
            player_b_name=match.competitor_b.profile.full_name if match.competitor_b and match.competitor_b.profile else None,
            player_a_avatar=match.competitor_a.effective_avatar_url if match.competitor_a else None,
            player_b_avatar=match.competitor_b.effective_avatar_url if match.competitor_b else None,
        )
        legs_dict[leg_key].append(match_data)

        if match.status == TAMatchStatus.COMPLETED:
            total_completed += 1
        elif match.status == TAMatchStatus.IN_PROGRESS:
            current_leg_key = leg_key

    # Define phase order for sorting
    phase_order = {
        TATournamentPhase.QUALIFIER.value: 0,
        TATournamentPhase.REQUALIFICATION.value: 1,
        TATournamentPhase.SEMIFINAL.value: 2,
        TATournamentPhase.FINAL_GRAND.value: 3,
        TATournamentPhase.FINAL_SMALL.value: 4,
    }

    # Sort legs by phase first, then by leg number
    sorted_keys = sorted(legs_dict.keys(), key=lambda k: (phase_order.get(k[1], 99), k[0]))

    # If no in-progress match, find the first incomplete leg
    if current_leg_key is None:
        for leg_key in sorted_keys:
            leg_matches = legs_dict[leg_key]
            if any(m.status != TAMatchStatusAPI.COMPLETED for m in leg_matches):
                current_leg_key = leg_key
                break

    # Build legs response - each entry is unique by (leg_number, phase)
    legs_list = []
    for leg_key in sorted_keys:
        leg_num, leg_phase = leg_key
        leg_matches = legs_dict[leg_key]
        completed_in_leg = sum(1 for m in leg_matches if m.status == TAMatchStatusAPI.COMPLETED)
        is_completed = completed_in_leg == len(leg_matches)
        legs_list.append(TARoundResponse(
            leg_number=leg_num,
            phase=TATournamentPhaseAPI(leg_phase) if leg_phase else TATournamentPhaseAPI.QUALIFIER,
            matches=leg_matches,
            matches_completed=completed_in_leg,
            total_matches=len(leg_matches),
            is_current=(leg_key == current_leg_key),
            is_completed=is_completed,
        ))

    # Extract leg number from key for backwards compatibility (just the number)
    current_leg_num = current_leg_key[0] if current_leg_key else None

    return {
        "legs": legs_list,
        "current_leg": current_leg_num,
        "total_legs": len(legs_list),
        "matches_completed": total_completed,
        "total_matches": len(matches),
        # Backwards compatibility aliases
        "rounds": legs_list,
        "current_round": current_leg_num,
        "total_rounds": len(legs_list),
    }


@router.get("/events/{event_id}/my-match", response_model=TAMatchResponse)
async def get_my_current_match(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> TAMatch:
    """
    Get the current user's active match in a TA event.

    Returns the user's current in-progress or pending match.
    Returns 404 if no active match found.
    """
    event = await get_ta_event(event_id, db, request)

    # Find user's active match (in_progress or pending)
    match_query = (
        select(TAMatch)
        .options(
            selectinload(TAMatch.competitor_a).selectinload(UserAccount.profile),
            selectinload(TAMatch.competitor_b).selectinload(UserAccount.profile),
        )
        .where(
            TAMatch.event_id == event_id,
            TAMatch.status.in_([TAMatchStatus.IN_PROGRESS.value, TAMatchStatus.SCHEDULED.value]),
            (TAMatch.competitor_a_id == current_user.id) | (TAMatch.competitor_b_id == current_user.id),
        )
        .order_by(TAMatch.round_number, TAMatch.match_number)
        .limit(1)
    )

    result = await db.execute(match_query)
    match = result.scalar_one_or_none()

    if not match:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No active match found",
        )

    # Return match with player info
    return TAMatchResponse(
        id=match.id,
        event_id=match.event_id,
        leg_number=match.round_number,  # Map to leg_number for API
        match_number=match.match_number,
        phase=TATournamentPhaseAPI(match.phase) if match.phase else TATournamentPhaseAPI.QUALIFIER,
        player_a_id=match.competitor_a_id,
        player_b_id=match.competitor_b_id,
        seat_a=match.seat_a,
        seat_b=match.seat_b,
        player_a_catches=match.competitor_a_catches or 0,
        player_b_catches=match.competitor_b_catches or 0,
        player_a_points=match.competitor_a_points or 0,
        player_b_points=match.competitor_b_points or 0,
        player_a_outcome=TAMatchOutcomeAPI(match.competitor_a_outcome_code) if match.competitor_a_outcome_code else None,
        player_b_outcome=TAMatchOutcomeAPI(match.competitor_b_outcome_code) if match.competitor_b_outcome_code else None,
        status=TAMatchStatusAPI(match.status) if match.status else TAMatchStatusAPI.PENDING,
        started_at=match.started_at,
        completed_at=match.completed_at,
        created_at=match.created_at,
        player_a_name=match.competitor_a.profile.full_name if match.competitor_a and match.competitor_a.profile else None,
        player_b_name=match.competitor_b.profile.full_name if match.competitor_b and match.competitor_b.profile else None,
        player_a_avatar=match.competitor_a.effective_avatar_url if match.competitor_a else None,
        player_b_avatar=match.competitor_b.effective_avatar_url if match.competitor_b else None,
    )


# =============================================================================
# Lineup Endpoints
# =============================================================================

@router.get("/events/{event_id}/lineups", response_model=TALineupListResponse)
async def list_lineups(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Get all lineups for a TA event."""
    event = await get_ta_event(event_id, db, request)

    query = (
        select(TALineup)
        .options(
            selectinload(TALineup.user).selectinload(UserAccount.profile),
            selectinload(TALineup.club),
        )
        .where(TALineup.event_id == event_id)
        .order_by(TALineup.draw_number)
    )
    result = await db.execute(query)
    lineups = result.scalars().all()

    # Build response with user info
    items = []
    has_ghost = False
    sectors = set()

    for lineup in lineups:
        if lineup.is_ghost:
            has_ghost = True

        sectors.add(lineup.sector)

        item = {
            "id": lineup.id,
            "event_id": lineup.event_id,
            "user_id": lineup.user_id,
            "enrollment_id": lineup.enrollment_id,
            "team_id": lineup.team_id,
            "club_id": lineup.club_id,
            "club_name": lineup.club.name if lineup.club else None,
            "draw_number": lineup.draw_number,
            "sector": lineup.sector,
            "initial_seat": lineup.initial_seat,
            "is_ghost": lineup.is_ghost,
            "created_at": lineup.created_at,
            "user_name": None,
            "user_avatar": None,
        }

        if lineup.user and lineup.user.profile:
            item["user_name"] = lineup.user.profile.full_name
            item["user_avatar"] = lineup.user.effective_avatar_url

        items.append(item)

    return {
        "items": items,
        "total": len(items),
        "has_ghost": has_ghost,
        "sectors": len(sectors),
    }


@router.get("/events/{event_id}/algorithm-preview", response_model=TAAlgorithmPreviewResponse)
async def get_algorithm_preview(
    event_id: int,
    request: Request,
    match_duration_minutes: int = Query(default=15, ge=5, le=60),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Preview available algorithms before generating lineups.

    Shows for each algorithm:
    - Number of legs (rounds)
    - Matches per leg
    - Total matches
    - Estimated duration
    - Warnings (e.g., "Very long for 40+ participants")

    Helps organizer choose the best algorithm for their event.
    """
    event = await get_ta_event(event_id, db, request, require_settings=False)

    # Get enrolled count
    enrolled_query = select(func.count()).select_from(EventEnrollment).where(
        EventEnrollment.event_id == event_id,
        EventEnrollment.status == "approved",
    )
    enrolled_result = await db.execute(enrolled_query)
    enrolled_count = enrolled_result.scalar() or 0

    if enrolled_count < 2:
        return {
            "event_id": event_id,
            "enrolled_count": enrolled_count,
            "effective_participants": 0,
            "has_ghost": False,
            "options": [],
            "recommended_algorithm": PairingAlgorithmAPI.ROUND_ROBIN_HALF,
        }

    # Calculate effective participants (add ghost if odd)
    has_ghost = enrolled_count % 2 == 1
    effective_participants = enrolled_count + (1 if has_ghost else 0)
    matches_per_leg = effective_participants // 2

    # Build algorithm options
    options = []

    # Half Round Robin (N/2 legs) - Recommended for most cases
    half_legs = effective_participants // 2
    half_total_matches = half_legs * matches_per_leg
    half_duration = TAPairingService.calculate_event_duration(
        num_participants=enrolled_count,
        algorithm=PairingAlgorithm.ROUND_ROBIN_HALF,
        match_duration_minutes=match_duration_minutes,
    )
    options.append(TAAlgorithmOption(
        algorithm=PairingAlgorithmAPI.ROUND_ROBIN_HALF,
        name="Half Round Robin",
        description="Each competitor plays half the field",
        legs=half_legs,
        matches_per_leg=matches_per_leg,
        total_matches=half_total_matches,
        estimated_duration_formatted=half_duration["total_duration_formatted"],
        is_recommended=enrolled_count <= 30,
        warning=None if enrolled_count <= 30 else "May be long for large groups",
    ))

    # Full Round Robin (N-1 legs)
    full_legs = effective_participants - 1
    full_total_matches = full_legs * matches_per_leg
    full_duration = TAPairingService.calculate_event_duration(
        num_participants=enrolled_count,
        algorithm=PairingAlgorithm.ROUND_ROBIN_FULL,
        match_duration_minutes=match_duration_minutes,
    )
    options.append(TAAlgorithmOption(
        algorithm=PairingAlgorithmAPI.ROUND_ROBIN_FULL,
        name="Full Round Robin",
        description="Everyone plays everyone",
        legs=full_legs,
        matches_per_leg=matches_per_leg,
        total_matches=full_total_matches,
        estimated_duration_formatted=full_duration["total_duration_formatted"],
        is_recommended=enrolled_count <= 12,
        warning="Very long for more than 12 participants" if enrolled_count > 12 else None,
    ))

    # Quarter Round Robin (N/4 legs) - for large groups
    quarter_legs = max(effective_participants // 4, 1)
    quarter_total_matches = quarter_legs * matches_per_leg
    quarter_duration = TAPairingService.calculate_event_duration(
        num_participants=enrolled_count,
        algorithm=PairingAlgorithm.ROUND_ROBIN_CUSTOM,
        match_duration_minutes=match_duration_minutes,
        custom_rounds=quarter_legs,
    )
    options.append(TAAlgorithmOption(
        algorithm=PairingAlgorithmAPI.ROUND_ROBIN_CUSTOM,
        name="Quarter Round Robin",
        description="Shorter tournament, N/4 legs",
        legs=quarter_legs,
        matches_per_leg=matches_per_leg,
        total_matches=quarter_total_matches,
        estimated_duration_formatted=quarter_duration["total_duration_formatted"],
        is_recommended=enrolled_count > 30,
        warning=None,
    ))

    # Simple Pairs (single leg)
    simple_duration = TAPairingService.calculate_event_duration(
        num_participants=enrolled_count,
        algorithm=PairingAlgorithm.SIMPLE_PAIRS,
        match_duration_minutes=match_duration_minutes,
    )
    options.append(TAAlgorithmOption(
        algorithm=PairingAlgorithmAPI.SIMPLE_PAIRS,
        name="Simple Pairs",
        description="Single round, quick tournament",
        legs=1,
        matches_per_leg=matches_per_leg,
        total_matches=matches_per_leg,
        estimated_duration_formatted=simple_duration["total_duration_formatted"],
        is_recommended=False,
        warning="Only one match per competitor",
    ))

    # Determine recommended algorithm
    if enrolled_count <= 12:
        recommended = PairingAlgorithmAPI.ROUND_ROBIN_FULL
    elif enrolled_count <= 30:
        recommended = PairingAlgorithmAPI.ROUND_ROBIN_HALF
    else:
        recommended = PairingAlgorithmAPI.ROUND_ROBIN_CUSTOM

    return {
        "event_id": event_id,
        "enrolled_count": enrolled_count,
        "effective_participants": effective_participants,
        "has_ghost": has_ghost,
        "options": options,
        "recommended_algorithm": recommended,
    }


@router.post(
    "/events/{event_id}/lineups/generate",
    response_model=TAGenerateLineupResponse,
)
async def generate_lineups(
    event_id: int,
    data: TAGenerateLineupRequest,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Generate lineups for a TA event using the specified pairing algorithm.

    This creates:
    1. Lineup entries with draw numbers, sectors, and initial seats
    2. Match schedule based on the pairing algorithm
    3. Game cards for each match

    Supports:
    - round_robin_full: Everyone plays everyone (N-1 rounds)
    - round_robin_half: Everyone plays half the field (N/2 rounds)
    - round_robin_custom: Specify exact number of rounds
    - simple_pairs: Single round n/2 pairing

    If participant count is odd, a ghost participant is added automatically.
    """
    event = await get_ta_event(event_id, db, request)

    # Guard: Block lineup generation for ongoing/completed events
    require_modifiable_status(event, action="generate lineups")

    settings = event.ta_settings

    # Delete existing lineups, matches, and game cards if regenerating
    # First delete game cards (depends on matches)
    await db.execute(
        TAGameCard.__table__.delete().where(TAGameCard.event_id == event_id)
    )
    # Delete standings
    await db.execute(
        TAQualifierStanding.__table__.delete().where(TAQualifierStanding.event_id == event_id)
    )
    # Delete matches
    await db.execute(
        TAMatch.__table__.delete().where(TAMatch.event_id == event_id)
    )
    # Delete lineups
    await db.execute(
        TALineup.__table__.delete().where(TALineup.event_id == event_id)
    )

    # Get enrolled participants
    enrollments_query = (
        select(EventEnrollment)
        .options(selectinload(EventEnrollment.user).selectinload(UserAccount.profile))
        .where(
            EventEnrollment.event_id == event_id,
            EventEnrollment.status == "approved",
        )
        .order_by(EventEnrollment.draw_number.nullslast(), EventEnrollment.id)
    )
    enrollments_result = await db.execute(enrollments_query)
    enrollments = enrollments_result.scalars().all()

    if len(enrollments) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=get_error_message("ta_not_enough_participants", request, min=2),
        )

    # Prepare participants for pairing service
    participants = []
    for enrollment in enrollments:
        name = f"Player {enrollment.user_id}"
        if enrollment.user and enrollment.user.profile:
            profile = enrollment.user.profile
            name = f"{profile.first_name} {profile.last_name}".strip() or name

        participants.append({
            "user_id": enrollment.user_id,
            "enrollment_id": enrollment.id,
            "name": name,
        })

    # Get club_ids for all participants (for club-based reporting)
    user_ids = [p["user_id"] for p in participants if p["user_id"]]
    club_lookup = await get_club_ids_for_users(db, user_ids)

    # Add club_id to each participant
    for p in participants:
        p["club_id"] = club_lookup.get(p["user_id"]) if p["user_id"] else None

    # ========================================================================
    # Use TAPairingService with Circle Method for proper round-robin pairings
    # This works correctly for all N values including N=4
    # ========================================================================
    import random

    # 1. Shuffle participants for random draw order
    random.shuffle(participants)

    # 2. Use TAPairingService to generate proper round-robin pairings
    algorithm = map_pairing_algorithm(data.algorithm)
    pairing_service = TAPairingService()
    pairing_result = pairing_service.generate_pairing(
        participants=participants,
        algorithm=algorithm,
        custom_rounds=data.custom_legs,
    )

    # Get values from pairing result
    N = pairing_result.total_participants
    has_ghost = pairing_result.has_ghost
    total_legs = pairing_result.total_rounds
    matches_per_leg = pairing_result.matches_per_round

    # Build participant lookup by name for later use
    # The pairing service assigns IDs 1-N to participants
    participant_lookup = {}
    for i, p in enumerate(participants):
        p["draw_number"] = i + 1
        p["is_ghost"] = False
        participant_lookup[f"P{i + 1}"] = p
        # Use the name from pairing service's participant list
        if i < len(pairing_service.participants):
            participant_lookup[pairing_service.participants[i].name] = p

    # Add ghost to lookup if present
    if has_ghost:
        ghost_participant = {
            "user_id": None,
            "enrollment_id": None,
            "name": "GHOST",
            "is_ghost": True,
            "draw_number": N,
        }
        participant_lookup["GHOST"] = ghost_participant
        participant_lookup[f"GHOST-{N}"] = ghost_participant
        # Also add by pairing service's ghost name
        if pairing_service.ghost:
            participant_lookup[pairing_service.ghost.name] = ghost_participant

    # Update enrollments with draw numbers
    for p in participants:
        if p.get("enrollment_id"):
            for enrollment in enrollments:
                if enrollment.id == p["enrollment_id"]:
                    enrollment.draw_number = p["draw_number"]
                    break

    # 3. Create lineups and matches from pairing result
    created_lineups = []

    for round_idx, round_matches in enumerate(pairing_result.rounds):
        leg = round_idx + 1

        # Track seat assignments for this leg (for lineup entries)
        seat_assignments = {}  # draw_number -> seat_number

        # Process each match in this round
        for match in round_matches:
            # Get participant info from lookup
            p_a_name = match.participant_a.name
            p_b_name = match.participant_b.name
            p_a = participant_lookup.get(p_a_name) or participant_lookup.get(repr(match.participant_a))
            p_b = participant_lookup.get(p_b_name) or participant_lookup.get(repr(match.participant_b))

            if p_a:
                seat_assignments[p_a.get("draw_number", match.seat_a)] = match.seat_a
            if p_b:
                seat_assignments[p_b.get("draw_number", match.seat_b)] = match.seat_b

        # Create lineup entries for all participants in this leg
        for draw_num in range(1, N + 1):
            seat_num = seat_assignments.get(draw_num, draw_num)

            # Find participant by draw number
            participant = None
            for p in participants:
                if p.get("draw_number") == draw_num:
                    participant = p
                    break

            # Check if this is the ghost
            is_ghost = participant is None or participant.get("is_ghost", False)
            if participant is None and has_ghost and draw_num == N:
                is_ghost = True

            lineup = TALineup(
                event_id=event_id,
                leg_number=leg,
                user_id=participant["user_id"] if participant and not is_ghost else None,
                enrollment_id=participant["enrollment_id"] if participant and not is_ghost else None,
                club_id=participant.get("club_id") if participant and not is_ghost else None,
                draw_number=draw_num,
                sector=1,
                seat_number=seat_num,
                is_ghost=is_ghost,
            )
            db.add(lineup)
            created_lineups.append(lineup)

        # Create matches from pairing_result (proper circle method pairings)
        for match_idx, pairing_match in enumerate(round_matches):
            match_num = match_idx + 1

            # Get participant info from pairing result
            p_a_name = pairing_match.participant_a.name
            p_b_name = pairing_match.participant_b.name
            p_a = participant_lookup.get(p_a_name) or participant_lookup.get(repr(pairing_match.participant_a))
            p_b = participant_lookup.get(p_b_name) or participant_lookup.get(repr(pairing_match.participant_b))

            # Handle case where participant not found (shouldn't happen)
            if p_a is None:
                p_a = {"user_id": None, "is_ghost": pairing_match.participant_a.is_ghost}
            if p_b is None:
                p_b = {"user_id": None, "is_ghost": pairing_match.participant_b.is_ghost}

            is_ghost_match = pairing_match.is_ghost_match
            ghost_side = pairing_match.ghost_side

            ta_match = TAMatch(
                event_id=event_id,
                phase=TATournamentPhase.QUALIFIER.value,
                leg_number=leg,
                round_number=leg,
                match_number=match_num,
                competitor_a_id=p_a.get("user_id"),
                competitor_b_id=p_b.get("user_id"),
                seat_a=pairing_match.seat_a,
                seat_b=pairing_match.seat_b,
                is_ghost_match=is_ghost_match,
                ghost_side=ghost_side,
                status=TAMatchStatus.IN_PROGRESS.value,
            )
            db.add(ta_match)
            await db.flush()

            # Create game cards
            if not p_a.get("is_ghost", False) and p_a.get("user_id"):
                card_a = TAGameCard(
                    event_id=event_id,
                    match_id=ta_match.id,
                    leg_number=leg,
                    user_id=p_a["user_id"],
                    my_seat=pairing_match.seat_a,
                    opponent_id=p_b.get("user_id"),
                    opponent_seat=pairing_match.seat_b if not p_b.get("is_ghost", False) else None,
                    is_ghost_opponent=p_b.get("is_ghost", False),
                    status=TAGameCardStatus.DRAFT.value,
                )
                db.add(card_a)

            if not p_b.get("is_ghost", False) and p_b.get("user_id"):
                card_b = TAGameCard(
                    event_id=event_id,
                    match_id=ta_match.id,
                    leg_number=leg,
                    user_id=p_b["user_id"],
                    my_seat=pairing_match.seat_b,
                    opponent_id=p_a.get("user_id"),
                    opponent_seat=pairing_match.seat_a if not p_a.get("is_ghost", False) else None,
                    is_ghost_opponent=p_a.get("is_ghost", False),
                    status=TAGameCardStatus.DRAFT.value,
                )
                db.add(card_b)

    # Get real participant count for response
    real_participants = pairing_result.real_participants
    total_matches = total_legs * matches_per_leg

    # Update settings with algorithm and draw info
    settings.pairing_algorithm = data.algorithm.value if hasattr(data.algorithm, 'value') else data.algorithm
    if data.custom_legs:
        settings.custom_legs = data.custom_legs
    settings.total_rounds = total_legs  # Keep model field name
    settings.matches_per_round = matches_per_leg  # Keep model field name
    settings.additional_rules = {
        **settings.additional_rules,
        "draw_completed": True,
        "total_legs": total_legs,
        "matches_per_leg": matches_per_leg,
        # Backwards compatibility
        "total_rounds": total_legs,
        "matches_per_round": matches_per_leg,
    }
    settings.updated_at = datetime.now(timezone.utc)

    await db.commit()

    # Calculate duration estimate
    duration = TAPairingService.calculate_event_duration(
        num_participants=len(participants),
        algorithm=algorithm,
        match_duration_minutes=settings.match_duration_minutes,
        custom_rounds=data.custom_legs,
    )

    # Refresh lineups for response
    for lineup in created_lineups:
        await db.refresh(lineup)

    # Build lineup response items
    lineup_items = []
    for lineup in created_lineups:
        item = TALineupResponse(
            id=lineup.id,
            event_id=lineup.event_id,
            user_id=lineup.user_id,
            enrollment_id=lineup.enrollment_id,
            team_id=lineup.team_id,
            club_id=lineup.club_id,
            club_name=lineup.club.name if lineup.club else None,
            draw_number=lineup.draw_number,
            sector=lineup.sector,
            initial_seat=lineup.initial_seat,
            is_ghost=lineup.is_ghost,
            created_at=lineup.created_at,
        )
        lineup_items.append(item)

    return {
        "message": get_error_message("lineup_created", request, count=real_participants),
        "total_participants": N,
        "real_participants": real_participants,
        "has_ghost": has_ghost,
        "algorithm": data.algorithm.value,
        "total_legs": total_legs,
        "matches_per_leg": matches_per_leg,
        "total_matches": total_matches,
        "estimated_duration": duration["total_duration_formatted"],
        "lineups": lineup_items,
    }


# =============================================================================
# Match Endpoints
# =============================================================================

@router.get("/events/{event_id}/matches", response_model=TAMatchListResponse)
async def list_matches(
    event_id: int,
    request: Request,
    phase: Optional[TATournamentPhaseAPI] = None,
    leg_number: Optional[int] = Query(None, alias="leg"),
    status_filter: Optional[str] = Query(None, alias="status"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    List matches for a TA event.

    Filters:
    - phase: Filter by tournament phase
    - leg: Filter by leg number (mansă)
    - status: Filter by match status
    """
    await get_ta_event(event_id, db, request)

    query = (
        select(TAMatch)
        .options(
            selectinload(TAMatch.competitor_a).selectinload(UserAccount.profile),
            selectinload(TAMatch.competitor_b).selectinload(UserAccount.profile),
        )
        .where(TAMatch.event_id == event_id)
    )

    if phase:
        query = query.where(TAMatch.phase == phase.value)
    if leg_number:
        query = query.where(TAMatch.round_number == leg_number)
    if status_filter:
        query = query.where(TAMatch.status == status_filter)

    query = query.order_by(TAMatch.round_number, TAMatch.match_number)

    result = await db.execute(query)
    matches = result.scalars().all()

    # Build response with user info
    items = []
    by_leg: dict[int, list] = {}

    for match in matches:
        # Get player names from loaded relationships
        player_a_name = None
        player_a_avatar = None
        if match.competitor_a and match.competitor_a.profile:
            player_a_name = match.competitor_a.profile.full_name
            player_a_avatar = match.competitor_a.effective_avatar_url

        player_b_name = None
        player_b_avatar = None
        if match.competitor_b and match.competitor_b.profile:
            player_b_name = match.competitor_b.profile.full_name
            player_b_avatar = match.competitor_b.effective_avatar_url

        item = TAMatchResponse(
            id=match.id,
            event_id=match.event_id,
            phase=TATournamentPhaseAPI(match.phase),
            leg_number=match.round_number,  # Map to leg_number for API
            match_number=match.match_number,
            player_a_id=match.competitor_a_id,
            player_b_id=match.competitor_b_id,
            seat_a=match.seat_a,
            seat_b=match.seat_b,
            player_a_catches=match.competitor_a_catches,
            player_b_catches=match.competitor_b_catches,
            player_a_points=match.competitor_a_points,
            player_b_points=match.competitor_b_points,
            player_a_outcome=TAMatchOutcome(match.competitor_a_outcome_code) if match.competitor_a_outcome_code else None,
            player_b_outcome=TAMatchOutcome(match.competitor_b_outcome_code) if match.competitor_b_outcome_code else None,
            status=match.status,
            started_at=match.started_at,
            completed_at=match.completed_at,
            created_at=match.created_at,
            player_a_name=player_a_name,
            player_b_name=player_b_name,
            player_a_avatar=player_a_avatar,
            player_b_avatar=player_b_avatar,
        )
        items.append(item)

        leg_num = match.round_number
        if leg_num not in by_leg:
            by_leg[leg_num] = []
        by_leg[leg_num].append(item)

    return {
        "items": items,
        "total": len(items),
        "by_leg": by_leg,
    }


# NOTE: This export route MUST be defined before /matches/{match_id} to avoid
# "export" being matched as a match_id by FastAPI's path parameter routing
@router.get("/events/{event_id}/matches/export")
async def export_matches_csv(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
):
    """
    Export TA match results to CSV.

    Includes all matches with competitor details, catches, points, and outcomes.
    Organized by leg and match number.
    """
    from fastapi.responses import StreamingResponse
    from io import StringIO
    import csv

    event = await get_ta_event(event_id, db, request, require_settings=False)
    event_name_safe = sanitize_filename(event.name) if event.name else f"Event_{event_id}"
    start_date_str = event.start_date.strftime("%d%m%Y") if event.start_date else "nodate"

    # Get all matches with competitor details
    matches_query = (
        select(TAMatch)
        .options(
            selectinload(TAMatch.competitor_a).selectinload(UserAccount.profile),
            selectinload(TAMatch.competitor_b).selectinload(UserAccount.profile),
        )
        .where(TAMatch.event_id == event_id)
        .order_by(TAMatch.leg_number, TAMatch.match_number)
    )
    result = await db.execute(matches_query)
    matches = result.scalars().all()

    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)

    # === SECTION: Event Info ===
    writer.writerow(["TA Match Results Export"])
    writer.writerow(["Event:", event.name or f"Event {event_id}"])
    writer.writerow(["Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    writer.writerow([])

    # === SECTION: Match Results ===
    writer.writerow(["=== MATCH RESULTS ==="])
    writer.writerow([])

    # Header
    writer.writerow([
        "Leg", "Match #", "Phase",
        "Player A Name", "Player A Draw #", "Player A Catches", "Player A Points", "Player A Outcome",
        "Player B Name", "Player B Draw #", "Player B Catches", "Player B Points", "Player B Outcome",
        "Status", "Is Ghost Match", "Completed At"
    ])

    # Data rows
    for match in matches:
        # Get player names
        player_a_name = ""
        if match.competitor_a and match.competitor_a.profile:
            player_a_name = f"{match.competitor_a.profile.last_name} {match.competitor_a.profile.first_name}".strip()
        elif match.competitor_a:
            player_a_name = f"User {match.competitor_a_id}"

        player_b_name = ""
        if match.competitor_b and match.competitor_b.profile:
            player_b_name = f"{match.competitor_b.profile.last_name} {match.competitor_b.profile.first_name}".strip()
        elif match.competitor_b:
            player_b_name = f"User {match.competitor_b_id}"

        # Handle ghost opponents
        if match.is_ghost_match:
            if match.ghost_side == "A":
                player_a_name = "[GHOST]"
            elif match.ghost_side == "B":
                player_b_name = "[GHOST]"

        writer.writerow([
            match.leg_number,
            match.match_number,
            match.phase or "qualifier",
            player_a_name,
            match.competitor_a_draw_number or "",
            match.competitor_a_catches if match.competitor_a_catches is not None else "",
            float(match.competitor_a_points) if match.competitor_a_points is not None else "",
            match.competitor_a_outcome_code or "",
            player_b_name,
            match.competitor_b_draw_number or "",
            match.competitor_b_catches if match.competitor_b_catches is not None else "",
            float(match.competitor_b_points) if match.competitor_b_points is not None else "",
            match.competitor_b_outcome_code or "",
            match.status or "",
            "Yes" if match.is_ghost_match else "No",
            match.completed_at.strftime("%Y-%m-%d %H:%M") if match.completed_at else "",
        ])

    writer.writerow([])

    # === SECTION: Summary Statistics ===
    writer.writerow(["=== SUMMARY ==="])
    writer.writerow([])

    total_matches = len(matches)
    completed_matches = sum(1 for m in matches if m.status == "completed")
    ghost_matches = sum(1 for m in matches if m.is_ghost_match)

    writer.writerow(["Total Matches:", total_matches])
    writer.writerow(["Completed Matches:", completed_matches])
    writer.writerow(["Ghost Matches:", ghost_matches])

    # Return CSV file
    output.seek(0)
    filename = f"{event_name_safe}_{start_date_str}_TA_Matches.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/events/{event_id}/matches/{match_id}", response_model=TAMatchDetailResponse)
async def get_match(
    event_id: int,
    match_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Get detailed match information including game cards."""
    await get_ta_event(event_id, db, request)

    query = (
        select(TAMatch)
        .options(
            selectinload(TAMatch.game_cards).selectinload(TAGameCard.user).selectinload(UserAccount.profile),
            selectinload(TAMatch.game_cards).selectinload(TAGameCard.opponent).selectinload(UserAccount.profile),
        )
        .where(TAMatch.id == match_id, TAMatch.event_id == event_id)
    )
    result = await db.execute(query)
    match = result.scalar_one_or_none()

    if not match:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=get_error_message("ta_match_not_found", request),
        )

    # Build game cards response (new per-user, per-leg structure)
    game_cards = []
    for card in match.game_cards:
        # Get points from match based on which player the card owner is
        my_points = None
        if card.is_validated:
            if card.user_id == match.competitor_a_id:
                my_points = float(match.competitor_a_points) if match.competitor_a_points else None
            elif card.user_id == match.competitor_b_id:
                my_points = float(match.competitor_b_points) if match.competitor_b_points else None

        game_cards.append(TAGameCardResponse(
            id=card.id,
            event_id=card.event_id,
            match_id=card.match_id,
            leg_number=card.leg_number,
            user_id=card.user_id,
            my_catches=card.my_catches,
            my_seat=card.my_seat,
            opponent_id=card.opponent_id,
            opponent_catches=card.opponent_catches,
            opponent_seat=card.opponent_seat,
            is_submitted=card.is_submitted,
            is_validated=card.is_validated,
            validated_at=card.validated_at,
            i_validated_opponent=card.i_validated_opponent,
            i_validated_at=card.i_validated_at,
            is_disputed=card.is_disputed,
            dispute_reason=card.dispute_reason,
            status=TAGameCardStatusAPI(card.status),
            is_ghost_opponent=card.is_ghost_opponent,
            submitted_at=card.submitted_at,
            created_at=card.created_at,
            updated_at=card.updated_at,
            my_points=my_points,
            user_name=card.user.profile.full_name if card.user and card.user.profile else None,
            user_avatar=card.user.effective_avatar_url if card.user else None,
            opponent_name=card.opponent.profile.full_name if card.opponent and card.opponent.profile else None,
            opponent_avatar=card.opponent.effective_avatar_url if card.opponent else None,
        ))

    return {
        "id": match.id,
        "event_id": match.event_id,
        "phase": TATournamentPhaseAPI(match.phase),
        "leg_number": match.round_number,  # Map to leg_number for API
        "match_number": match.match_number,
        "player_a_id": match.competitor_a_id,
        "player_b_id": match.competitor_b_id,
        "seat_a": match.seat_a,
        "seat_b": match.seat_b,
        "player_a_catches": match.competitor_a_catches or 0,
        "player_b_catches": match.competitor_b_catches or 0,
        "player_a_points": float(match.competitor_a_points) if match.competitor_a_points else 0,
        "player_b_points": float(match.competitor_b_points) if match.competitor_b_points else 0,
        "player_a_outcome": TAMatchOutcomeAPI(match.competitor_a_outcome_code) if match.competitor_a_outcome_code else None,
        "player_b_outcome": TAMatchOutcomeAPI(match.competitor_b_outcome_code) if match.competitor_b_outcome_code else None,
        "status": TAMatchStatusAPI(match.status),
        "started_at": match.started_at,
        "completed_at": match.completed_at,
        "created_at": match.created_at,
        "game_cards": game_cards,
    }


@router.post("/events/{event_id}/matches/{match_id}/start", response_model=TAMatchResponse)
async def start_match(
    event_id: int,
    match_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> TAMatch:
    """Start a scheduled match."""
    await get_ta_event(event_id, db, request)

    query = select(TAMatch).where(TAMatch.id == match_id, TAMatch.event_id == event_id)
    result = await db.execute(query)
    match = result.scalar_one_or_none()

    if not match:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=get_error_message("ta_match_not_found", request),
        )

    if match.status != TAMatchStatus.SCHEDULED.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Match is not scheduled (current status: {match.status})",
        )

    match.status = TAMatchStatus.IN_PROGRESS.value
    match.started_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(match)

    return match


@router.post("/events/{event_id}/matches/start-leg")
async def start_leg_matches(
    event_id: int,
    leg_number: int = Query(..., description="Leg number to start"),
    phase: Optional[str] = Query(None, description="Phase filter (e.g., semifinal, final_grand)"),
    request: Request = None,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Start all scheduled matches in a specific leg/phase.

    This allows organizers to start an entire knockout round at once,
    enabling mobile users to input their scores.

    Args:
        event_id: Event ID
        leg_number: The display leg number (round_number in DB)
        phase: Optional phase filter (required for knockout phases)

    Returns:
        Count of matches started
    """
    await get_ta_event(event_id, db, request)

    # Find all scheduled matches for this leg (using round_number for display consistency)
    query = (
        select(TAMatch)
        .where(
            TAMatch.event_id == event_id,
            TAMatch.round_number == leg_number,
            TAMatch.status == TAMatchStatus.SCHEDULED.value,
        )
    )

    if phase:
        query = query.where(TAMatch.phase == phase)

    result = await db.execute(query)
    matches = result.scalars().all()

    if not matches:
        return {
            "message": "No scheduled matches found for this leg",
            "started_count": 0,
        }

    # Start all matches
    now = datetime.now(timezone.utc)
    for match in matches:
        match.status = TAMatchStatus.IN_PROGRESS.value
        match.started_at = now

    await db.commit()

    return {
        "message": f"Started {len(matches)} matches",
        "started_count": len(matches),
        "leg_number": leg_number,
        "phase": phase,
    }


@router.patch("/events/{event_id}/matches/{match_id}/results", response_model=TAMatchResponse)
async def edit_match_results(
    event_id: int,
    match_id: int,
    data: TAMatchResultUpdate,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Edit match results (catches for competitors A and/or B).

    Only organizers and admins can edit match results.
    Edit history is tracked with edited_by_id, edited_at, and previous values.
    After editing, outcomes and points are recalculated using event's point config.
    """
    event = await get_ta_event(event_id, db, request)

    query = select(TAMatch).where(TAMatch.id == match_id, TAMatch.event_id == event_id)
    result = await db.execute(query)
    match = result.scalar_one_or_none()

    if not match:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=get_error_message("ta_match_not_found", request),
        )

    # Get point config for this event (custom or defaults)
    point_config_query = select(TAEventPointConfig).where(TAEventPointConfig.event_id == event_id)
    point_result = await db.execute(point_config_query)
    point_config = point_result.scalar_one_or_none()

    # Store previous values for audit
    update_data = data.model_dump(exclude_unset=True)
    if 'competitor_a_catches' in update_data:
        match.previous_a_catches = match.competitor_a_catches
    if 'competitor_b_catches' in update_data:
        match.previous_b_catches = match.competitor_b_catches

    # Track who edited and when
    match.edited_by_id = current_user.id
    match.edited_at = datetime.now(timezone.utc)

    # Apply updates
    for key, value in update_data.items():
        if hasattr(match, key):
            setattr(match, key, value)

    # Recalculate outcomes and points if both catches are set
    if match.competitor_a_catches is not None and match.competitor_b_catches is not None:
        match.calculate_outcome(point_config)  # This modifies the match object in place
        match.status = TAMatchStatus.COMPLETED.value
        if not match.completed_at:
            match.completed_at = datetime.now(timezone.utc)

        # Update both game cards to reflect admin entry
        # Get both cards for this match
        cards_query = select(TAGameCard).where(TAGameCard.match_id == match_id)
        cards_result = await db.execute(cards_query)
        cards = cards_result.scalars().all()

        now = datetime.now(timezone.utc)
        for card in cards:
            # Set catches based on which player this card belongs to
            if card.user_id == match.competitor_a_id:
                card.my_catches = match.competitor_a_catches
                card.opponent_catches = match.competitor_b_catches
            elif card.user_id == match.competitor_b_id:
                card.my_catches = match.competitor_b_catches
                card.opponent_catches = match.competitor_a_catches

            # Mark as submitted and fully validated by admin
            card.is_submitted = True
            if not card.submitted_at:
                card.submitted_at = now
            card.is_validated = True
            if not card.validated_at:
                card.validated_at = now
                card.validated_by_id = current_user.id
            card.i_validated_opponent = True
            if not card.i_validated_at:
                card.i_validated_at = now

            # Update status to validated (fully completed)
            card.status = TAGameCardStatus.VALIDATED.value
            card.updated_at = now

        # Cascade updates to downstream phases if this is a knockout match
        cascade_result = await _cascade_knockout_update(db, event_id, match)

        # Broadcast SSE match result for live public leaderboard
        winner_id = None
        if match.competitor_a_catches > match.competitor_b_catches:
            winner_id = match.competitor_a_id
        elif match.competitor_b_catches > match.competitor_a_catches:
            winner_id = match.competitor_b_id

        await broadcast_ta_match_result(
            event_id=event_id,
            match_id=match.id,
            phase=match.phase,
            leg_number=match.round_number,
            competitor_a_id=match.competitor_a_id,
            competitor_b_id=match.competitor_b_id,
            competitor_a_catches=match.competitor_a_catches or 0,
            competitor_b_catches=match.competitor_b_catches or 0,
            winner_id=winner_id,
        )

        # Trigger stats recalculation for both players
        if match.competitor_a_id:
            await statistics_service.update_user_stats_for_event(db, match.competitor_a_id, event_id)
        if match.competitor_b_id:
            await statistics_service.update_user_stats_for_event(db, match.competitor_b_id, event_id)

    await db.commit()
    await db.refresh(match)

    # Build response with correct field mappings
    return {
        "id": match.id,
        "event_id": match.event_id,
        "phase": TATournamentPhaseAPI(match.phase),
        "leg_number": match.round_number,  # Map to leg_number for API
        "match_number": match.match_number,
        "player_a_id": match.competitor_a_id,
        "player_b_id": match.competitor_b_id,
        "seat_a": match.seat_a,
        "seat_b": match.seat_b,
        "player_a_catches": match.competitor_a_catches or 0,
        "player_b_catches": match.competitor_b_catches or 0,
        "player_a_points": match.competitor_a_points or Decimal("0.0"),
        "player_b_points": match.competitor_b_points or Decimal("0.0"),
        "player_a_outcome": TAMatchOutcomeAPI(match.competitor_a_outcome_code) if match.competitor_a_outcome_code else None,
        "player_b_outcome": TAMatchOutcomeAPI(match.competitor_b_outcome_code) if match.competitor_b_outcome_code else None,
        "status": TAMatchStatusAPI(match.status),
        "started_at": match.started_at,
        "completed_at": match.completed_at,
        "created_at": match.created_at,
    }


# =============================================================================
# Game Card Endpoints
# =============================================================================

@router.get("/events/{event_id}/game-cards/my", response_model=TAMyGameCardsResponse)
async def get_my_game_cards(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Get current user's game cards for a TA event (all legs)."""
    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    query = (
        select(TAGameCard)
        .options(
            selectinload(TAGameCard.user).selectinload(UserAccount.profile),
            selectinload(TAGameCard.opponent).selectinload(UserAccount.profile),
            selectinload(TAGameCard.match),  # Load match to get points
        )
        .where(
            TAGameCard.event_id == event_id,
            TAGameCard.user_id == current_user.id,
        )
        .order_by(TAGameCard.leg_number)
    )
    result = await db.execute(query)
    cards = result.scalars().all()

    # Determine current leg (first non-validated leg)
    current_leg = None
    for card in cards:
        if not card.is_validated:
            current_leg = card.leg_number
            break

    items = []
    for card in cards:
        # Get points from match based on which player the card owner is
        my_points = None
        if card.match and card.is_validated:
            if card.user_id == card.match.competitor_a_id:
                my_points = float(card.match.competitor_a_points) if card.match.competitor_a_points else None
            elif card.user_id == card.match.competitor_b_id:
                my_points = float(card.match.competitor_b_points) if card.match.competitor_b_points else None

        items.append(TAGameCardResponse(
            id=card.id,
            event_id=card.event_id,
            match_id=card.match_id,
            leg_number=card.leg_number,
            phase=TATournamentPhaseAPI(card.match.phase) if card.match and card.match.phase else TATournamentPhaseAPI.QUALIFIER,
            user_id=card.user_id,
            my_catches=card.my_catches,
            my_seat=card.my_seat,
            opponent_id=card.opponent_id,
            opponent_catches=card.opponent_catches,
            opponent_seat=card.opponent_seat,
            is_submitted=card.is_submitted,
            is_validated=card.is_validated,
            validated_at=card.validated_at,
            i_validated_opponent=card.i_validated_opponent,
            i_validated_at=card.i_validated_at,
            is_disputed=card.is_disputed,
            dispute_reason=card.dispute_reason,
            status=TAGameCardStatusAPI(card.status),
            is_ghost_opponent=card.is_ghost_opponent,
            submitted_at=card.submitted_at,
            created_at=card.created_at,
            updated_at=card.updated_at,
            my_points=my_points,
            user_name=card.user.profile.full_name if card.user and card.user.profile else None,
            user_avatar=card.user.effective_avatar_url if card.user else None,
            opponent_name=card.opponent.profile.full_name if card.opponent and card.opponent.profile else None,
            opponent_avatar=card.opponent.effective_avatar_url if card.opponent else None,
        ))

    return {
        "items": items,
        "total": len(items),
        "current_leg": current_leg,
        "event_id": event_id,
    }


@router.post(
    "/events/{event_id}/game-cards/{card_id}/submit",
    response_model=TAGameCardResponse,
)
async def submit_game_card(
    event_id: int,
    card_id: int,
    data: TAGameCardSubmitRequest,
    request: Request,
    current_user: UserAccount = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Submit a game card with catches count.

    Self-validation flow:
    1. User enters their catches
    2. Card is marked as submitted
    3. If opponent already submitted, both see each other's catches
    4. If ghost opponent, auto-validates immediately

    The card can only be submitted by its owner.
    """
    await get_ta_event(event_id, db, request)

    # Get game card with row-level lock to prevent concurrent updates
    query = (
        select(TAGameCard)
        .options(
            selectinload(TAGameCard.user).selectinload(UserAccount.profile),
            selectinload(TAGameCard.opponent).selectinload(UserAccount.profile),
        )
        .where(
            TAGameCard.id == card_id,
            TAGameCard.event_id == event_id,
        )
        .with_for_update(nowait=False)  # Row-level pessimistic lock
    )
    result = await db.execute(query)
    card = result.scalar_one_or_none()

    if not card:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=get_error_message("ta_game_card_not_found", request),
        )

    # Verify ownership
    if card.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=get_error_message("ta_participant_not_in_match", request),
        )

    # Check status
    if card.status not in [TAGameCardStatus.DRAFT.value, TAGameCardStatus.DISPUTED.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=get_error_message("ta_game_card_locked", request),
        )

    # Update card with catches
    card.my_catches = data.my_catches
    card.is_submitted = True
    card.status = TAGameCardStatus.SUBMITTED.value
    card.submitted_at = datetime.now(timezone.utc)
    card.updated_at = datetime.now(timezone.utc)

    # If ghost opponent, auto-validate (both directions) and complete match
    if card.is_ghost_opponent:
        card.opponent_catches = 0
        # Ghost validates my catches (auto)
        card.is_validated = True
        card.validated_at = datetime.now(timezone.utc)
        # I validate ghost's catches (auto - they have 0)
        card.i_validated_opponent = True
        card.i_validated_at = datetime.now(timezone.utc)
        card.status = TAGameCardStatus.VALIDATED.value

        # Sync catches to TAMatch and complete it (BYE match auto-completion)
        if card.match_id:
            match_query = select(TAMatch).where(TAMatch.id == card.match_id)
            match_result = await db.execute(match_query)
            match = match_result.scalar_one_or_none()
            if match:
                # Set catches on match - user is competitor_a or competitor_b
                if match.competitor_a_id == card.user_id:
                    match.competitor_a_catches = card.my_catches
                    match.competitor_b_catches = 0  # Ghost always 0
                else:
                    match.competitor_b_catches = card.my_catches
                    match.competitor_a_catches = 0  # Ghost always 0

                # Calculate outcome using point config
                point_config_query = select(TAEventPointConfig).where(
                    TAEventPointConfig.event_id == event_id
                )
                point_result = await db.execute(point_config_query)
                point_config = point_result.scalar_one_or_none()
                match.calculate_outcome(point_config)

                # Mark match as completed
                match.status = TAMatchStatus.COMPLETED.value
                match.completed_at = datetime.now(timezone.utc)

                # Auto-update standings for ghost match
                await update_standings_for_match(db, match, point_config)

                # Trigger stats recalculation for the player (ghost has no user_id)
                if card.user_id:
                    await statistics_service.update_user_stats_for_event(db, card.user_id, event_id)

    # Check if opponent has already submitted - if so, update opponent_catches
    if card.opponent_id:
        opponent_card_query = select(TAGameCard).where(
            TAGameCard.event_id == event_id,
            TAGameCard.leg_number == card.leg_number,
            TAGameCard.user_id == card.opponent_id,
        )
        opponent_result = await db.execute(opponent_card_query)
        opponent_card = opponent_result.scalar_one_or_none()

        if opponent_card and opponent_card.is_submitted:
            # Both have submitted - update both cards with opponent catches
            card.opponent_catches = opponent_card.my_catches
            opponent_card.opponent_catches = card.my_catches
            opponent_card.updated_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(card)

    # Broadcast SSE event to notify opponent
    try:
        await redis_cache.publish_sse_event(event_id, {
            "type": "game_card_submitted",
            "match_id": card.match_id,
            "leg_number": card.leg_number,
            "user_id": card.user_id,
            "opponent_id": card.opponent_id,
        })
    except Exception:
        # Don't fail the request if SSE broadcast fails
        pass

    return {
        "id": card.id,
        "event_id": card.event_id,
        "match_id": card.match_id,
        "leg_number": card.leg_number,
        "user_id": card.user_id,
        "my_catches": card.my_catches,
        "my_seat": card.my_seat,
        "opponent_id": card.opponent_id,
        "opponent_catches": card.opponent_catches,
        "opponent_seat": card.opponent_seat,
        "is_submitted": card.is_submitted,
        "is_validated": card.is_validated,
        "validated_at": card.validated_at,
        "i_validated_opponent": card.i_validated_opponent,
        "i_validated_at": card.i_validated_at,
        "is_disputed": card.is_disputed,
        "dispute_reason": card.dispute_reason,
        "status": TAGameCardStatusAPI(card.status),
        "is_ghost_opponent": card.is_ghost_opponent,
        "submitted_at": card.submitted_at,
        "created_at": card.created_at,
        "updated_at": card.updated_at,
        "user_name": card.user.profile.full_name if card.user and card.user.profile else None,
        "user_avatar": card.user.effective_avatar_url if card.user else None,
        "opponent_name": card.opponent.profile.full_name if card.opponent and card.opponent.profile else None,
    }


@router.post(
    "/events/{event_id}/game-cards/{card_id}/validate",
    response_model=TAGameCardResponse,
)
async def validate_opponent_card(
    event_id: int,
    card_id: int,
    data: TAGameCardValidateRequest,
    request: Request,
    current_user: UserAccount = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Validate opponent's game card (self-validation in TA).

    Self-validation flow:
    1. Opponent submits their catches
    2. This user validates opponent's card
    3. If valid, card is marked validated
    4. If disputed, admin must resolve
    5. When both cards in a match are validated, match is complete

    This endpoint allows a participant to validate their opponent's card.
    """
    event = await get_ta_event(event_id, db, request)

    # Get the opponent's game card with row-level lock to prevent concurrent updates
    query = (
        select(TAGameCard)
        .options(
            selectinload(TAGameCard.user).selectinload(UserAccount.profile),
            selectinload(TAGameCard.opponent).selectinload(UserAccount.profile),
            selectinload(TAGameCard.match),
        )
        .where(
            TAGameCard.id == card_id,
            TAGameCard.event_id == event_id,
        )
        .with_for_update(nowait=False)  # Row-level pessimistic lock
    )
    result = await db.execute(query)
    opponent_card = result.scalar_one_or_none()

    if not opponent_card:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=get_error_message("ta_game_card_not_found", request),
        )

    # Cannot validate own card
    if opponent_card.user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=get_error_message("ta_cannot_validate_own_card", request),
        )

    # Must be the opponent in the match
    if opponent_card.opponent_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=get_error_message("ta_participant_not_in_match", request),
        )

    # Card must be submitted
    if not opponent_card.is_submitted:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Card must be submitted before validation",
        )

    # Already validated
    if opponent_card.is_validated:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=get_error_message("ta_already_validated", request),
        )

    # Get my card to update i_validated_opponent
    my_card_query = select(TAGameCard).where(
        TAGameCard.event_id == event_id,
        TAGameCard.leg_number == opponent_card.leg_number,
        TAGameCard.user_id == current_user.id,
    )
    my_card_result = await db.execute(my_card_query)
    my_card = my_card_result.scalar_one_or_none()

    if data.is_valid:
        # Mark opponent's card as validated by me
        opponent_card.is_validated = True
        opponent_card.validated_by_id = current_user.id
        opponent_card.validated_at = datetime.now(timezone.utc)
        opponent_card.status = TAGameCardStatus.VALIDATED.value

        # Mark my card as having validated opponent
        if my_card:
            my_card.i_validated_opponent = True
            my_card.i_validated_at = datetime.now(timezone.utc)
            my_card.updated_at = datetime.now(timezone.utc)

        # Check if BOTH cards are now validated - if so, update match result
        if my_card and my_card.is_validated:
            # Both validated - update match results
            match = opponent_card.match
            if match:
                # Determine which side is which
                if match.competitor_a_id == opponent_card.user_id:
                    match.competitor_a_catches = opponent_card.my_catches
                    match.competitor_b_catches = my_card.my_catches
                else:
                    match.competitor_b_catches = opponent_card.my_catches
                    match.competitor_a_catches = my_card.my_catches

                # Sync opponent_catches on both game cards to match final values
                opponent_card.opponent_catches = my_card.my_catches
                my_card.opponent_catches = opponent_card.my_catches
                my_card.updated_at = datetime.now(timezone.utc)

                # Calculate outcome
                point_config_query = select(TAEventPointConfig).where(TAEventPointConfig.event_id == event_id)
                point_result = await db.execute(point_config_query)
                point_config = point_result.scalar_one_or_none()
                match.calculate_outcome(point_config)

                match.status = TAMatchStatus.COMPLETED.value
                match.completed_at = datetime.now(timezone.utc)

                # Auto-update standings when match completes
                await update_standings_for_match(db, match, point_config)

                # Trigger stats recalculation for both players
                if match.competitor_a_id:
                    await statistics_service.update_user_stats_for_event(db, match.competitor_a_id, event_id)
                if match.competitor_b_id:
                    await statistics_service.update_user_stats_for_event(db, match.competitor_b_id, event_id)
    else:
        opponent_card.is_disputed = True
        opponent_card.dispute_reason = data.dispute_reason
        opponent_card.status = TAGameCardStatus.DISPUTED.value

    opponent_card.updated_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(opponent_card)

    # Broadcast SSE event to notify the card owner (whose card was validated)
    try:
        await redis_cache.publish_sse_event(event_id, {
            "type": "game_card_validated",
            "match_id": opponent_card.match_id,
            "leg_number": opponent_card.leg_number,
            "card_id": opponent_card.id,
            "user_id": opponent_card.user_id,
            "validated_by": current_user.id,
            "is_disputed": opponent_card.is_disputed,
        })
    except Exception:
        # Don't fail the request if SSE broadcast fails
        pass

    return {
        "id": opponent_card.id,
        "event_id": opponent_card.event_id,
        "match_id": opponent_card.match_id,
        "leg_number": opponent_card.leg_number,
        "user_id": opponent_card.user_id,
        "my_catches": opponent_card.my_catches,
        "my_seat": opponent_card.my_seat,
        "opponent_id": opponent_card.opponent_id,
        "opponent_catches": opponent_card.opponent_catches,
        "opponent_seat": opponent_card.opponent_seat,
        "is_submitted": opponent_card.is_submitted,
        "is_validated": opponent_card.is_validated,
        "validated_at": opponent_card.validated_at,
        "i_validated_opponent": opponent_card.i_validated_opponent,
        "i_validated_at": opponent_card.i_validated_at,
        "is_disputed": opponent_card.is_disputed,
        "dispute_reason": opponent_card.dispute_reason,
        "status": TAGameCardStatusAPI(opponent_card.status),
        "is_ghost_opponent": opponent_card.is_ghost_opponent,
        "submitted_at": opponent_card.submitted_at,
        "created_at": opponent_card.created_at,
        "updated_at": opponent_card.updated_at,
        "user_name": opponent_card.user.profile.full_name if opponent_card.user and opponent_card.user.profile else None,
        "user_avatar": opponent_card.user.effective_avatar_url if opponent_card.user else None,
        "opponent_name": opponent_card.opponent.profile.full_name if opponent_card.opponent and opponent_card.opponent.profile else None,
    }


# =============================================================================
# Admin Game Card Endpoints (Validator/Organizer)
# =============================================================================

@router.get("/events/{event_id}/game-cards", response_model=TAMyGameCardsResponse)
async def get_user_game_cards(
    event_id: int,
    request: Request,
    user_id: Optional[int] = Query(None, description="User ID to fetch cards for"),
    current_user: UserAccount = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get game cards for a specific user (admin/validator access).

    If user_id is not provided, returns cards for current user.
    Requires validator/organizer/admin permissions to fetch other users' cards.
    """
    event = await get_ta_event(event_id, db, request)

    target_user_id = user_id if user_id else current_user.id

    # If fetching another user's cards, check permissions
    if target_user_id != current_user.id:
        # Check if current user is organizer, validator, or admin
        user_roles = set(current_user.profile.roles or []) if current_user.profile else set()
        is_admin = "administrator" in user_roles
        is_organizer = "organizer" in user_roles
        is_event_owner = event.created_by_id == current_user.id

        # Also allow if current user is enrolled in the same event (public competition data)
        is_participant = False
        if not (is_admin or is_organizer or is_event_owner):
            enrollment_check = await db.execute(
                select(EventEnrollment.id).where(
                    EventEnrollment.event_id == event_id,
                    EventEnrollment.user_id == current_user.id,
                    EventEnrollment.status == "approved",
                )
            )
            is_participant = enrollment_check.scalar() is not None

        if not (is_admin or is_organizer or is_event_owner or is_participant):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not authorized to view other users' game cards",
            )

    query = (
        select(TAGameCard)
        .options(
            selectinload(TAGameCard.user).selectinload(UserAccount.profile),
            selectinload(TAGameCard.opponent).selectinload(UserAccount.profile),
            selectinload(TAGameCard.match),  # Load match to get points
        )
        .where(
            TAGameCard.event_id == event_id,
            TAGameCard.user_id == target_user_id,
        )
        .order_by(TAGameCard.leg_number)
    )
    result = await db.execute(query)
    cards = result.scalars().all()

    # Determine current leg (first non-validated leg)
    current_leg = None
    for card in cards:
        if not card.is_validated:
            current_leg = card.leg_number
            break

    items = []
    for card in cards:
        # Get points from match based on which player the card owner is
        my_points = None
        if card.match and card.is_validated:
            if card.user_id == card.match.competitor_a_id:
                my_points = float(card.match.competitor_a_points) if card.match.competitor_a_points else None
            elif card.user_id == card.match.competitor_b_id:
                my_points = float(card.match.competitor_b_points) if card.match.competitor_b_points else None

        items.append(TAGameCardResponse(
            id=card.id,
            event_id=card.event_id,
            match_id=card.match_id,
            leg_number=card.leg_number,
            phase=TATournamentPhaseAPI(card.match.phase) if card.match and card.match.phase else TATournamentPhaseAPI.QUALIFIER,
            user_id=card.user_id,
            my_catches=card.my_catches,
            my_seat=card.my_seat,
            opponent_id=card.opponent_id,
            opponent_catches=card.opponent_catches,
            opponent_seat=card.opponent_seat,
            is_submitted=card.is_submitted,
            is_validated=card.is_validated,
            validated_at=card.validated_at,
            i_validated_opponent=card.i_validated_opponent,
            i_validated_at=card.i_validated_at,
            is_disputed=card.is_disputed,
            dispute_reason=card.dispute_reason,
            status=TAGameCardStatusAPI(card.status),
            is_ghost_opponent=card.is_ghost_opponent,
            submitted_at=card.submitted_at,
            created_at=card.created_at,
            updated_at=card.updated_at,
            my_points=my_points,
            user_name=card.user.profile.full_name if card.user and card.user.profile else None,
            user_avatar=card.user.effective_avatar_url if card.user else None,
            opponent_name=card.opponent.profile.full_name if card.opponent and card.opponent.profile else None,
            opponent_avatar=card.opponent.effective_avatar_url if card.opponent else None,
        ))

    return {
        "items": items,
        "total": len(items),
        "current_leg": current_leg,
        "event_id": event_id,
    }


@router.patch(
    "/events/{event_id}/game-cards/{card_id}/admin-update",
    response_model=TAGameCardResponse,
)
async def admin_update_game_card(
    event_id: int,
    card_id: int,
    data: TAGameCardAdminUpdateRequest,
    request: Request,
    current_user: UserAccount = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Admin/Validator update of a game card.

    Allows organizers and validators to manually update game card data
    for situations where users cannot submit themselves.
    """
    event = await get_ta_event(event_id, db, request)

    # Check permissions - must be organizer, validator, or admin
    user_roles = set(current_user.profile.roles or []) if current_user.profile else set()
    is_admin = "administrator" in user_roles
    is_organizer = "organizer" in user_roles
    is_validator = "validator" in user_roles
    is_event_owner = event.created_by_id == current_user.id

    if not (is_admin or is_organizer or is_validator or is_event_owner):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to update game cards",
        )

    # Get the card
    query = (
        select(TAGameCard)
        .options(
            selectinload(TAGameCard.user).selectinload(UserAccount.profile),
            selectinload(TAGameCard.opponent).selectinload(UserAccount.profile),
            selectinload(TAGameCard.match),
        )
        .where(
            TAGameCard.id == card_id,
            TAGameCard.event_id == event_id,
        )
    )
    result = await db.execute(query)
    card = result.scalar_one_or_none()

    if not card:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=get_error_message("ta_game_card_not_found", request),
        )

    # Update fields
    if data.my_catches is not None:
        card.my_catches = data.my_catches
        # Sync opponent's card's opponent_catches to keep data consistent
        if card.opponent_id:
            opp_card_sync_query = select(TAGameCard).where(
                TAGameCard.match_id == card.match_id,
                TAGameCard.user_id == card.opponent_id,
            )
            opp_result_sync = await db.execute(opp_card_sync_query)
            opp_card_sync = opp_result_sync.scalar_one_or_none()
            if opp_card_sync:
                opp_card_sync.opponent_catches = data.my_catches
                opp_card_sync.updated_at = datetime.now(timezone.utc)
    if data.is_submitted is not None:
        card.is_submitted = data.is_submitted
        if data.is_submitted and not card.submitted_at:
            card.submitted_at = datetime.now(timezone.utc)
    if data.is_validated is not None:
        card.is_validated = data.is_validated
        if data.is_validated and not card.validated_at:
            card.validated_at = datetime.now(timezone.utc)
            card.validated_by_id = current_user.id
    if data.i_validated_opponent is not None:
        card.i_validated_opponent = data.i_validated_opponent
        if data.i_validated_opponent and not card.i_validated_at:
            card.i_validated_at = datetime.now(timezone.utc)

    # For BYE matches (ghost opponent), auto-validate opponent since there is none
    if card.is_ghost_opponent and card.is_submitted:
        card.i_validated_opponent = True
        if not card.i_validated_at:
            card.i_validated_at = datetime.now(timezone.utc)
        # Also set opponent catches to 0 for BYE
        card.opponent_catches = 0

    card.updated_at = datetime.now(timezone.utc)

    # Update status based on new state
    if card.is_disputed:
        card.status = TAGameCardStatus.DISPUTED.value
    elif card.is_validated and card.i_validated_opponent:
        card.status = TAGameCardStatus.VALIDATED.value
    elif card.is_validated:
        card.status = TAGameCardStatus.VALIDATED.value
    elif card.is_submitted:
        card.status = TAGameCardStatus.SUBMITTED.value

    # Check if match should be completed
    if card.match and card.is_validated and card.i_validated_opponent:
        match = card.match

        # Handle BYE matches (ghost opponent) - complete immediately
        if card.is_ghost_opponent:
            # For BYE: user wins by default with their catches vs 0
            if match.player_a_id == card.user_id:
                match.player_a_catches = card.my_catches
                match.player_b_catches = 0
            else:
                match.player_b_catches = card.my_catches
                match.player_a_catches = 0
            match.status = TAMatchStatus.COMPLETED.value
            match.completed_at = datetime.now(timezone.utc)

            # Get point config and calculate outcomes
            point_config_query = select(TAEventPointConfig).where(
                TAEventPointConfig.event_id == card.event_id
            )
            point_result = await db.execute(point_config_query)
            point_config = point_result.scalar_one_or_none()
            match.calculate_outcome(point_config)

            # Auto-update standings when match completes
            await update_standings_for_match(db, match, point_config)

            # Trigger stats recalculation for the player (ghost has no user_id)
            if card.user_id:
                await statistics_service.update_user_stats_for_event(db, card.user_id, card.event_id)
        else:
            # Regular match - check opponent's card
            opp_card_query = select(TAGameCard).where(
                TAGameCard.match_id == card.match_id,
                TAGameCard.user_id == card.opponent_id,
            )
            opp_result = await db.execute(opp_card_query)
            opp_card = opp_result.scalar_one_or_none()

            if opp_card and opp_card.is_validated and opp_card.i_validated_opponent:
                # Both cards validated, update match
                match.player_a_catches = card.my_catches if match.player_a_id == card.user_id else opp_card.my_catches
                match.player_b_catches = opp_card.my_catches if match.player_a_id == card.user_id else card.my_catches
                match.status = TAMatchStatus.COMPLETED.value
                match.completed_at = datetime.now(timezone.utc)

                # Get point config and calculate outcomes
                point_config_query = select(TAEventPointConfig).where(
                    TAEventPointConfig.event_id == card.event_id
                )
                point_result = await db.execute(point_config_query)
                point_config = point_result.scalar_one_or_none()
                match.calculate_outcome(point_config)

                # Auto-update standings when match completes
                await update_standings_for_match(db, match, point_config)

                # Trigger stats recalculation for both players
                if match.player_a_id:
                    await statistics_service.update_user_stats_for_event(db, match.player_a_id, card.event_id)
                if match.player_b_id:
                    await statistics_service.update_user_stats_for_event(db, match.player_b_id, card.event_id)

    await db.commit()
    await db.refresh(card)

    return {
        "id": card.id,
        "event_id": card.event_id,
        "match_id": card.match_id,
        "leg_number": card.leg_number,
        "user_id": card.user_id,
        "my_catches": card.my_catches,
        "my_seat": card.my_seat,
        "opponent_id": card.opponent_id,
        "opponent_catches": card.opponent_catches,
        "opponent_seat": card.opponent_seat,
        "is_submitted": card.is_submitted,
        "is_validated": card.is_validated,
        "validated_at": card.validated_at,
        "i_validated_opponent": card.i_validated_opponent,
        "i_validated_at": card.i_validated_at,
        "is_disputed": card.is_disputed,
        "dispute_reason": card.dispute_reason,
        "status": TAGameCardStatusAPI(card.status),
        "is_ghost_opponent": card.is_ghost_opponent,
        "submitted_at": card.submitted_at,
        "created_at": card.created_at,
        "updated_at": card.updated_at,
        "user_name": card.user.profile.full_name if card.user and card.user.profile else None,
        "user_avatar": card.user.effective_avatar_url if card.user else None,
        "opponent_name": card.opponent.profile.full_name if card.opponent and card.opponent.profile else None,
    }


# =============================================================================
# Standings Endpoints
# =============================================================================

@router.post("/events/{event_id}/standings/recalculate", response_model=MessageResponse)
async def recalculate_standings(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Recalculate all standings for a TA event.

    This rebuilds the qualifier standings table from all completed matches.
    Useful after manual match edits or data corrections.
    """
    from app.services.ta_ranking import TARankingService

    event = await get_ta_event(event_id, db, request)

    # Get all completed matches
    matches_query = select(TAMatch).where(
        TAMatch.event_id == event_id,
        TAMatch.status == TAMatchStatus.COMPLETED.value,
    )
    result = await db.execute(matches_query)
    matches = result.scalars().all()

    # Get user->enrollment mapping
    from app.models.event import EventEnrollment
    enrollments_query = select(EventEnrollment).where(
        EventEnrollment.event_id == event_id,
        EventEnrollment.status == "approved",
    )
    result = await db.execute(enrollments_query)
    enrollments = result.scalars().all()
    user_enrollment_map = {e.user_id: e.id for e in enrollments}

    # Clear existing standings
    await db.execute(
        TAQualifierStanding.__table__.delete().where(TAQualifierStanding.event_id == event_id)
    )

    # Rebuild standings from matches
    ranking_service = TARankingService(db)

    # Accumulate stats by user
    user_stats: dict[int, dict] = {}

    for match in matches:
        if match.competitor_a_id:
            if match.competitor_a_id not in user_stats:
                user_stats[match.competitor_a_id] = {
                    "total_points": Decimal("0"),
                    "total_fish_caught": 0,
                    "total_matches": 0,
                    "total_victories": 0,
                    "total_ties": 0,
                    "total_losses": 0,
                }
            stats = user_stats[match.competitor_a_id]
            stats["total_points"] += match.competitor_a_points or Decimal("0")
            stats["total_fish_caught"] += match.competitor_a_catches or 0
            stats["total_matches"] += 1
            if match.competitor_a_outcome_code == "V":
                stats["total_victories"] += 1
            elif match.competitor_a_outcome_code in ["T", "T0"]:
                stats["total_ties"] += 1
            else:
                stats["total_losses"] += 1

        if match.competitor_b_id:
            if match.competitor_b_id not in user_stats:
                user_stats[match.competitor_b_id] = {
                    "total_points": Decimal("0"),
                    "total_fish_caught": 0,
                    "total_matches": 0,
                    "total_victories": 0,
                    "total_ties": 0,
                    "total_losses": 0,
                }
            stats = user_stats[match.competitor_b_id]
            stats["total_points"] += match.competitor_b_points or Decimal("0")
            stats["total_fish_caught"] += match.competitor_b_catches or 0
            stats["total_matches"] += 1
            if match.competitor_b_outcome_code == "V":
                stats["total_victories"] += 1
            elif match.competitor_b_outcome_code in ["T", "T0"]:
                stats["total_ties"] += 1
            else:
                stats["total_losses"] += 1

    # Create standing records
    for user_id, stats in user_stats.items():
        enrollment_id = user_enrollment_map.get(user_id)
        if not enrollment_id:
            continue  # Skip users without enrollment

        standing = TAQualifierStanding(
            event_id=event_id,
            user_id=user_id,
            enrollment_id=enrollment_id,
            rank=0,  # Will be calculated
            total_points=stats["total_points"],
            total_fish_caught=stats["total_fish_caught"],
            total_matches=stats["total_matches"],
            total_victories=stats["total_victories"],
            total_ties=stats["total_ties"],
            total_losses=stats["total_losses"],
        )
        db.add(standing)

    await db.flush()

    # Recalculate ranks
    await ranking_service._recalculate_ranks(event_id)
    await db.commit()

    # Broadcast SSE standings update for live public leaderboard
    await broadcast_ta_standings_update(
        event_id=event_id,
        phase=TATournamentPhase.QUALIFIER.value,
        top_changes=None,  # Full recalculation doesn't track individual changes
    )

    return {"message": f"Standings recalculated for {len(user_stats)} participants", "details": {"participants_ranked": len(user_stats)}}


@router.post(
    "/events/{event_id}/generate-bracket",
    response_model=TAGenerateBracketResponse,
)
async def generate_knockout_bracket(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Generate knockout bracket after qualifier phase is complete.

    This creates matches for requalification (if enabled), semifinals, and finals.
    Game cards are created for each match.

    Requirements:
    - All qualifier matches must be completed
    - Event must have knockout stage enabled

    Process:
    1. Verify all qualifier matches are completed
    2. Get final standings from qualifier
    3. Create bracket based on settings:
       - If requalification: positions N+1 to N+M compete for extra spots
       - Semifinals: Top 4 (or top 2 + requalification winners)
       - Finals: Grand final (1st/2nd) and Small final (3rd/4th)
    4. Create game cards for each bracket match
    """
    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    if not settings:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Event has no TA settings configured",
        )

    if not settings.has_knockout_stage:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Event does not have knockout stage enabled",
        )

    # Check if all qualifier matches are completed
    qualifier_matches_query = select(TAMatch).where(
        TAMatch.event_id == event_id,
        TAMatch.phase == TATournamentPhase.QUALIFIER.value,
    )
    result = await db.execute(qualifier_matches_query)
    qualifier_matches = result.scalars().all()

    if not qualifier_matches:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No qualifier matches found. Generate lineups first.",
        )

    incomplete_matches = [m for m in qualifier_matches if m.status != TAMatchStatus.COMPLETED.value]
    if incomplete_matches:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot generate bracket: {len(incomplete_matches)} qualifier matches not completed",
        )

    # Check if bracket already exists
    existing_bracket_query = select(TAKnockoutBracket).where(
        TAKnockoutBracket.event_id == event_id
    )
    existing_result = await db.execute(existing_bracket_query)
    existing_bracket = existing_result.scalar_one_or_none()

    if existing_bracket:
        # Delete existing bracket and its matches/cards
        await db.execute(
            TAGameCard.__table__.delete().where(
                TAGameCard.event_id == event_id,
                TAGameCard.match_id.in_(
                    select(TAMatch.id).where(
                        TAMatch.event_id == event_id,
                        TAMatch.phase != TATournamentPhase.QUALIFIER.value,
                    )
                )
            )
        )
        await db.execute(
            TAMatch.__table__.delete().where(
                TAMatch.event_id == event_id,
                TAMatch.phase != TATournamentPhase.QUALIFIER.value,
            )
        )
        await db.delete(existing_bracket)
        await db.flush()

    # Get standings sorted by rank
    standings_query = (
        select(TAQualifierStanding)
        .options(selectinload(TAQualifierStanding.user).selectinload(UserAccount.profile))
        .where(TAQualifierStanding.event_id == event_id)
        .order_by(TAQualifierStanding.rank)
    )
    standings_result = await db.execute(standings_query)
    standings = standings_result.scalars().all()

    if len(standings) < 4:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Need at least 4 participants for knockout bracket",
        )

    # Build seeds map
    seeds = {}
    for standing in standings:
        seeds[str(standing.rank)] = standing.user_id

    # Create bracket record
    bracket = TAKnockoutBracket(
        event_id=event_id,
        total_qualifiers=settings.knockout_qualifiers,
        seeds=seeds,
        is_generated=True,
    )
    db.add(bracket)
    await db.flush()

    # Determine next leg number (after qualifier legs)
    max_leg_query = select(func.max(TAMatch.leg_number)).where(
        TAMatch.event_id == event_id,
        TAMatch.phase == TATournamentPhase.QUALIFIER.value,
    )
    max_leg_result = await db.execute(max_leg_query)
    max_qualifier_leg = max_leg_result.scalar() or 0
    next_leg = max_qualifier_leg + 1

    matches_created = []
    now = datetime.now(timezone.utc)

    # Helper to create match and game cards
    async def create_bracket_match(
        phase: str,
        match_number: int,
        leg_number: int,
        competitor_a_id: Optional[int],
        competitor_b_id: Optional[int],
        winner_placement: Optional[int] = None,
        loser_placement: Optional[int] = None,
    ) -> TAMatch:
        match = TAMatch(
            event_id=event_id,
            phase=phase,
            leg_number=leg_number,
            round_number=1,
            match_number=match_number,
            seat_a=1,  # Knockout matches use seat 1 (no seat rotation)
            seat_b=1,
            competitor_a_id=competitor_a_id,
            competitor_b_id=competitor_b_id,
            status=TAMatchStatus.SCHEDULED.value,
            is_ghost_match=False,
        )
        db.add(match)
        await db.flush()

        # Create game cards for both competitors
        if competitor_a_id:
            card_a = TAGameCard(
                event_id=event_id,
                match_id=match.id,
                leg_number=leg_number,
                user_id=competitor_a_id,
                opponent_id=competitor_b_id,
                my_seat=1,  # Knockout matches use seat 1
                opponent_seat=1,
                is_ghost_opponent=competitor_b_id is None,
                status=TAGameCardStatus.DRAFT.value,
            )
            db.add(card_a)

        if competitor_b_id:
            card_b = TAGameCard(
                event_id=event_id,
                match_id=match.id,
                leg_number=leg_number,
                user_id=competitor_b_id,
                opponent_id=competitor_a_id,
                my_seat=1,  # Knockout matches use seat 1
                opponent_seat=1,
                is_ghost_opponent=competitor_a_id is None,
                status=TAGameCardStatus.DRAFT.value,
            )
            db.add(card_b)

        return match

    # ========================================================================
    # REQUALIFICATION PHASE (if enabled)
    # ========================================================================
    requalification_winners = []
    if settings.has_requalification and settings.requalification_slots > 0:
        # Use direct_to_semifinal to determine who competes in requalification
        # E.g., if direct_to_semifinal=2 and requalification_slots=4:
        # Positions 3,4,5,6 compete: Match1: 3 vs 5, Match2: 4 vs 6
        # (Standard bracket: higher seed vs mid-range, not top vs bottom)
        direct_count = getattr(settings, 'direct_to_semifinal', 2)
        start_pos = direct_count + 1
        end_pos = direct_count + settings.requalification_slots

        requalification_positions = []
        for standing in standings:
            if start_pos <= standing.rank <= end_pos:
                requalification_positions.append(standing)

        # Create requalification matches (standard bracket seeding: 3v5, 4v6)
        # This ensures winners face the direct qualifiers fairly in semifinals:
        # - Winner of 3v5 faces seed 2 in semifinal
        # - Winner of 4v6 faces seed 1 in semifinal
        num_requalification = len(requalification_positions) // 2
        for i in range(num_requalification):
            # Match i pairs position i with position i+num_requalification
            # For 4 players [3,4,5,6]: Match1=3v5, Match2=4v6
            seed_a = requalification_positions[i] if i < len(requalification_positions) else None
            seed_b = requalification_positions[i + num_requalification] if (i + num_requalification) < len(requalification_positions) else None

            if seed_a and seed_b:
                match = await create_bracket_match(
                    phase=TATournamentPhase.REQUALIFICATION.value,
                    match_number=i + 1,
                    leg_number=next_leg,
                    competitor_a_id=seed_a.user_id,
                    competitor_b_id=seed_b.user_id,
                )
                matches_created.append(match)

        next_leg += 1

    # ========================================================================
    # SEMIFINALS
    # ========================================================================
    # Top 4 from qualifier (or top 2 + 2 requalification winners)
    semifinal_competitors = []

    if settings.has_requalification and settings.requalification_slots > 0:
        # Use direct_to_semifinal config (how many bypass requalification)
        direct_count = getattr(settings, 'direct_to_semifinal', 2)
        for standing in standings[:direct_count]:
            semifinal_competitors.append(standing.user_id)
        # Requalification winners will be determined later
        # For now, fill remaining semifinal slots with next ranked from qualifier
        # (requalification winners will replace these positions when they complete)
        requalification_winners = settings.requalification_slots // 2
        remaining_slots = requalification_winners
        for standing in standings[direct_count:direct_count + remaining_slots]:
            semifinal_competitors.append(standing.user_id)
    else:
        # Top 4 go directly to semifinals
        for standing in standings[:4]:
            semifinal_competitors.append(standing.user_id)

    # Semifinal bracket structure:
    # - SF1: Seed 1 vs Winner of Requalification Match 1 (3v5)
    # - SF2: Seed 2 vs Winner of Requalification Match 2 (4v6)
    # When requalification exists, positions 2 and 3 in semifinal_competitors
    # are placeholders that will be replaced by requalification winners.
    if len(semifinal_competitors) >= 4:
        sf1 = await create_bracket_match(
            phase=TATournamentPhase.SEMIFINAL.value,
            match_number=1,
            leg_number=next_leg,
            competitor_a_id=semifinal_competitors[0],  # Seed 1
            competitor_b_id=semifinal_competitors[2],  # Winner of 3v5 (placeholder: seed 3)
        )
        matches_created.append(sf1)

        sf2 = await create_bracket_match(
            phase=TATournamentPhase.SEMIFINAL.value,
            match_number=2,
            leg_number=next_leg,
            competitor_a_id=semifinal_competitors[1],  # Seed 2
            competitor_b_id=semifinal_competitors[3],  # Winner of 4v6 (placeholder: seed 4)
        )
        matches_created.append(sf2)

    next_leg += 1

    # ========================================================================
    # FINALS (Grand Final + Small Final)
    # ========================================================================
    # Finals competitors will be determined from semifinal results
    # Winners go to Grand Final (1st/2nd), Losers go to Small Final (3rd/4th)

    # Grand Final (1st/2nd place) - winners of semifinals
    grand_final = await create_bracket_match(
        phase=TATournamentPhase.FINAL_GRAND.value,
        match_number=1,
        leg_number=next_leg,
        competitor_a_id=None,  # Will be set when SF1 completes
        competitor_b_id=None,  # Will be set when SF2 completes
        winner_placement=1,
        loser_placement=2,
    )
    matches_created.append(grand_final)

    # Small Final (3rd/4th place) - losers of semifinals
    small_final = await create_bracket_match(
        phase=TATournamentPhase.FINAL_SMALL.value,
        match_number=1,
        leg_number=next_leg,
        competitor_a_id=None,  # Will be set when SF1 completes
        competitor_b_id=None,  # Will be set when SF2 completes
        winner_placement=3,
        loser_placement=4,
    )
    matches_created.append(small_final)

    # Set direct placements for those who didn't make knockout
    direct_placements = {}
    direct_count = getattr(settings, 'direct_to_semifinal', 2)
    if settings.has_requalification:
        # Those after requalification participants get direct placement
        placement_start = direct_count + settings.requalification_slots + 1
    else:
        # Top 4 in semifinals, rest get direct placement from 5th onwards
        placement_start = 5

    for standing in standings:
        if standing.rank >= placement_start:
            direct_placements[str(standing.rank)] = standing.user_id

    bracket.direct_placements = direct_placements

    await db.commit()

    # Broadcast SSE event for live public leaderboard
    semifinalist_ids = []
    requalification_ids = []
    for m in matches_created:
        if m.phase == TATournamentPhase.SEMIFINAL.value:
            if m.competitor_a_id:
                semifinalist_ids.append(m.competitor_a_id)
            if m.competitor_b_id:
                semifinalist_ids.append(m.competitor_b_id)
        elif m.phase == TATournamentPhase.REQUALIFICATION.value:
            if m.competitor_a_id:
                requalification_ids.append(m.competitor_a_id)
            if m.competitor_b_id:
                requalification_ids.append(m.competitor_b_id)

    await broadcast_ta_bracket_generated(
        event_id=event_id,
        semifinalists=list(set(semifinalist_ids)),
        requalification_participants=list(set(requalification_ids)) if requalification_ids else None,
    )
    await broadcast_ta_phase_advanced(
        event_id=event_id,
        from_phase=TATournamentPhase.QUALIFIER.value,
        to_phase=TATournamentPhase.SEMIFINAL.value if not settings.has_requalification else TATournamentPhase.REQUALIFICATION.value,
    )

    return {
        "message": f"Knockout bracket generated: {len(matches_created)} matches created",
        "matches_created": len(matches_created),
        "has_requalification": settings.has_requalification,
    }


@router.post(
    "/events/{event_id}/advance-to-finals",
)
async def advance_to_finals(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Advance semifinals winners/losers to finals.

    This endpoint:
    1. Checks that all semifinals are completed
    2. Determines winners and losers from each semifinal
    3. Populates Grand Final with semifinal winners
    4. Populates Small Final (3rd/4th) with semifinal losers
    5. Creates game cards for finals matches

    Requirements:
    - Both semifinal matches must be completed
    - Finals matches must exist (bracket already generated)
    """
    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    if not settings or not settings.has_knockout_stage:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Event does not have knockout stage enabled",
        )

    # Get semifinal matches
    semifinal_query = select(TAMatch).where(
        TAMatch.event_id == event_id,
        TAMatch.phase == TATournamentPhase.SEMIFINAL.value,
    ).order_by(TAMatch.match_number)
    result = await db.execute(semifinal_query)
    semifinals = result.scalars().all()

    if len(semifinals) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Semifinals not found. Generate bracket first.",
        )

    # Check all semifinals are completed
    incomplete = [m for m in semifinals if m.status != TAMatchStatus.COMPLETED.value]
    if incomplete:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot advance: {len(incomplete)} semifinal(s) not completed",
        )

    # Get finals matches
    finals_query = select(TAMatch).where(
        TAMatch.event_id == event_id,
        TAMatch.phase.in_([
            TATournamentPhase.FINAL_GRAND.value,
            TATournamentPhase.FINAL_SMALL.value,
        ]),
    )
    finals_result = await db.execute(finals_query)
    finals = {m.phase: m for m in finals_result.scalars().all()}

    if not finals:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Finals matches not found. Generate bracket first.",
        )

    # Determine winners and losers from semifinals
    sf_results = []
    for sf in semifinals:
        a_catches = sf.competitor_a_catches or 0
        b_catches = sf.competitor_b_catches or 0
        if a_catches > b_catches:
            winner_id = sf.competitor_a_id
            loser_id = sf.competitor_b_id
        elif b_catches > a_catches:
            winner_id = sf.competitor_b_id
            loser_id = sf.competitor_a_id
        else:
            # Tie - use tiebreaker (outcome codes)
            a_outcome = sf.competitor_a_outcome_code or ""
            b_outcome = sf.competitor_b_outcome_code or ""
            # V > T > T0 > L > L0 - but in a tie, use original seeding
            # For now, use competitor_a as winner on tie (higher seed)
            winner_id = sf.competitor_a_id
            loser_id = sf.competitor_b_id
        sf_results.append({"winner": winner_id, "loser": loser_id})

    # Update Grand Final with winners
    grand_final = finals.get(TATournamentPhase.FINAL_GRAND.value)
    if grand_final:
        grand_final.competitor_a_id = sf_results[0]["winner"]
        grand_final.competitor_b_id = sf_results[1]["winner"]

        # Create game cards for grand final competitors
        for user_id, opponent_id in [
            (sf_results[0]["winner"], sf_results[1]["winner"]),
            (sf_results[1]["winner"], sf_results[0]["winner"]),
        ]:
            existing_card = await db.execute(
                select(TAGameCard).where(
                    TAGameCard.match_id == grand_final.id,
                    TAGameCard.user_id == user_id,
                )
            )
            if not existing_card.scalar_one_or_none():
                card = TAGameCard(
                    event_id=event_id,
                    match_id=grand_final.id,
                    leg_number=grand_final.leg_number,
                    user_id=user_id,
                    opponent_id=opponent_id,
                    my_seat=1,
                    opponent_seat=1,
                    is_ghost_opponent=False,
                    status=TAGameCardStatus.DRAFT.value,
                )
                db.add(card)

    # Update Small Final with losers
    small_final = finals.get(TATournamentPhase.FINAL_SMALL.value)
    if small_final:
        small_final.competitor_a_id = sf_results[0]["loser"]
        small_final.competitor_b_id = sf_results[1]["loser"]

        # Create game cards for small final competitors
        for user_id, opponent_id in [
            (sf_results[0]["loser"], sf_results[1]["loser"]),
            (sf_results[1]["loser"], sf_results[0]["loser"]),
        ]:
            existing_card = await db.execute(
                select(TAGameCard).where(
                    TAGameCard.match_id == small_final.id,
                    TAGameCard.user_id == user_id,
                )
            )
            if not existing_card.scalar_one_or_none():
                card = TAGameCard(
                    event_id=event_id,
                    match_id=small_final.id,
                    leg_number=small_final.leg_number,
                    user_id=user_id,
                    opponent_id=opponent_id,
                    my_seat=1,
                    opponent_seat=1,
                    is_ghost_opponent=False,
                    status=TAGameCardStatus.DRAFT.value,
                )
                db.add(card)

    await db.commit()

    return {
        "message": "Advanced to finals successfully",
        "grand_final": {
            "competitor_a_id": sf_results[0]["winner"],
            "competitor_b_id": sf_results[1]["winner"],
        },
        "small_final": {
            "competitor_a_id": sf_results[0]["loser"],
            "competitor_b_id": sf_results[1]["loser"],
        },
    }


@router.post(
    "/events/{event_id}/advance-requalification",
)
async def advance_requalification_to_semifinals(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Advance requalification winners to semifinals.

    This endpoint:
    1. Checks that all requalification matches are completed
    2. Determines winners from each requalification match
    3. Updates semifinal matches with requalification winners (replacing placeholder positions)
    4. Updates game cards for semifinals with new opponent info

    Seeding logic:
    - Semifinal 1: Qualifier #1 vs Requalification Winner #2 (lower seed wins = harder path)
    - Semifinal 2: Qualifier #2 vs Requalification Winner #1 (higher seed wins = easier path)

    Requirements:
    - All requalification matches must be completed
    - Semifinal matches must exist (bracket already generated)
    - has_requalification must be enabled
    """
    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    if not settings or not settings.has_knockout_stage:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Event does not have knockout stage enabled",
        )

    if not settings.has_requalification:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Event does not have requalification enabled",
        )

    # Get requalification matches
    requalification_query = select(TAMatch).where(
        TAMatch.event_id == event_id,
        TAMatch.phase == TATournamentPhase.REQUALIFICATION.value,
    ).order_by(TAMatch.match_number)
    result = await db.execute(requalification_query)
    requalification_matches = result.scalars().all()

    if not requalification_matches:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Requalification matches not found. Generate bracket first.",
        )

    # Check all requalification matches are completed
    incomplete = [m for m in requalification_matches if m.status != TAMatchStatus.COMPLETED.value]
    if incomplete:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot advance: {len(incomplete)} requalification match(es) not completed",
        )

    # Determine winners from requalification
    requalification_winners = []
    for match in requalification_matches:
        a_catches = match.competitor_a_catches or 0
        b_catches = match.competitor_b_catches or 0
        if a_catches > b_catches:
            winner_id = match.competitor_a_id
        elif b_catches > a_catches:
            winner_id = match.competitor_b_id
        else:
            # Tie - use higher seed (competitor_a is always higher seed)
            winner_id = match.competitor_a_id
        requalification_winners.append(winner_id)

    # Get semifinal matches
    semifinal_query = select(TAMatch).where(
        TAMatch.event_id == event_id,
        TAMatch.phase == TATournamentPhase.SEMIFINAL.value,
    ).order_by(TAMatch.match_number)
    sf_result = await db.execute(semifinal_query)
    semifinals = list(sf_result.scalars().all())

    if len(semifinals) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Semifinals not found. Generate bracket first.",
        )

    # Update semifinals with requalification winners
    # Standard bracket seeding:
    # - SF1: Seed 1 vs Winner of 3v5 (requalification match #1)
    # - SF2: Seed 2 vs Winner of 4v6 (requalification match #2)

    old_sf1_b = semifinals[0].competitor_b_id
    old_sf2_b = semifinals[1].competitor_b_id

    if len(requalification_winners) >= 2:
        # SF1: Seed 1 vs Winner of 3v5 (requalification match #1)
        semifinals[0].competitor_b_id = requalification_winners[0]
        # SF2: Seed 2 vs Winner of 4v6 (requalification match #2)
        semifinals[1].competitor_b_id = requalification_winners[1]
    elif len(requalification_winners) == 1:
        # Only one requalification match - winner goes to SF1
        semifinals[0].competitor_b_id = requalification_winners[0]

    # Update game cards for affected semifinals
    for sf in semifinals:
        # Delete old game cards for the old competitor_b (placeholder)
        old_competitor = old_sf1_b if sf == semifinals[0] else old_sf2_b
        new_competitor = sf.competitor_b_id

        if old_competitor and old_competitor != new_competitor:
            # Delete old game card for removed competitor
            await db.execute(
                delete(TAGameCard).where(
                    TAGameCard.match_id == sf.id,
                    TAGameCard.user_id == old_competitor,
                )
            )

            # Update existing game card for competitor_a to point to new opponent
            await db.execute(
                update(TAGameCard).where(
                    TAGameCard.match_id == sf.id,
                    TAGameCard.user_id == sf.competitor_a_id,
                ).values(opponent_id=new_competitor)
            )

            # Create new game card for requalification winner
            existing_card = await db.execute(
                select(TAGameCard).where(
                    TAGameCard.match_id == sf.id,
                    TAGameCard.user_id == new_competitor,
                )
            )
            if not existing_card.scalar_one_or_none():
                card = TAGameCard(
                    event_id=event_id,
                    match_id=sf.id,
                    leg_number=sf.leg_number,
                    user_id=new_competitor,
                    opponent_id=sf.competitor_a_id,
                    my_seat=1,
                    opponent_seat=1,
                    is_ghost_opponent=False,
                    status=TAGameCardStatus.DRAFT.value,
                    phase=TATournamentPhase.SEMIFINAL.value,
                )
                db.add(card)

    await db.commit()

    return {
        "message": "Requalification winners advanced to semifinals",
        "requalification_winners": requalification_winners,
        "semifinal_1": {
            "competitor_a_id": semifinals[0].competitor_a_id,
            "competitor_b_id": semifinals[0].competitor_b_id,
        },
        "semifinal_2": {
            "competitor_a_id": semifinals[1].competitor_a_id,
            "competitor_b_id": semifinals[1].competitor_b_id,
        },
    }


@router.get(
    "/events/{event_id}/standings",
    response_model=TAQualifierStandingListResponse,
)
async def get_standings(
    event_id: int,
    request: Request,
    phase: Optional[TATournamentPhaseAPI] = None,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Get current standings for a TA event.

    For qualifier phase (or no phase specified): Returns stored TAQualifierStanding records.
    For knockout phases: Dynamically calculates standings from completed matches in that phase.
    """
    from app.services.ta_ranking import TARankingService

    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    items = []

    # Knockout phases that need dynamic calculation from matches
    knockout_phases = {
        TATournamentPhaseAPI.REQUALIFICATION,
        TATournamentPhaseAPI.SEMIFINAL,
        TATournamentPhaseAPI.FINAL_GRAND,
        TATournamentPhaseAPI.FINAL_SMALL,
    }

    if phase and phase in knockout_phases:
        # For knockout phases, calculate standings from completed matches
        ranking_service = TARankingService(db)
        rankings = await ranking_service.compute_leg_ranking(event_id, phase=phase.value)

        # Story 12.6: Get DQ status for all users in this phase
        user_ids = [r["user_id"] for r in rankings]
        dq_query = (
            select(EventEnrollment.user_id)
            .where(
                EventEnrollment.event_id == event_id,
                EventEnrollment.user_id.in_(user_ids),
                EventEnrollment.status == EnrollmentStatus.DISQUALIFIED.value,
            )
        )
        dq_result = await db.execute(dq_query)
        dq_user_ids = {row[0] for row in dq_result.fetchall()}

        dq_items = []
        non_dq_items = []

        for ranking in rankings:
            user_id = ranking["user_id"]
            is_dq = user_id in dq_user_ids

            item = TAQualifierStandingResponse(
                id=0,  # Dynamic calculation, no stored ID
                event_id=event_id,
                user_id=user_id,
                rank=None if is_dq else ranking.get("rank", 0),  # Story 12.6: No rank for DQ
                total_points=ranking["points"],
                total_catches=ranking["captures"],
                total_length=0.0,
                matches_played=ranking["matches_played"],
                victories=ranking["victories"],
                ties=ranking["ties_with_fish"] + ranking["ties_without_fish"],
                losses=ranking["losses_with_fish"] + ranking["losses_without_fish"],
                updated_at=datetime.now(timezone.utc),
                user_name=ranking.get("user_name"),
                user_avatar=ranking.get("user_avatar"),
                is_disqualified=is_dq,  # Story 12.6
            )

            if is_dq:
                dq_items.append(item)
            else:
                non_dq_items.append(item)

        # Story 12.6: DQ users appear at bottom
        items = non_dq_items + dq_items
    else:
        # For qualifier phase (or no phase), use stored TAQualifierStanding
        # Story 12.6: Include enrollment for DQ status
        query = (
            select(TAQualifierStanding)
            .options(
                selectinload(TAQualifierStanding.user).selectinload(UserAccount.profile),
                selectinload(TAQualifierStanding.enrollment),
            )
            .where(TAQualifierStanding.event_id == event_id)
            .order_by(TAQualifierStanding.rank)
        )
        result = await db.execute(query)
        standings = result.scalars().all()

        # Story 12.6: Separate DQ and non-DQ users
        dq_items = []
        non_dq_items = []

        for standing in standings:
            # Check DQ status from enrollment (Story 12.6)
            is_dq = (
                standing.enrollment is not None
                and standing.enrollment.status == EnrollmentStatus.DISQUALIFIED.value
            )

            item = TAQualifierStandingResponse(
                id=standing.id,
                event_id=standing.event_id,
                user_id=standing.user_id,
                rank=None if is_dq else standing.rank,  # Story 12.6: No rank for DQ users
                total_points=standing.total_points,
                total_catches=standing.total_fish_caught,
                total_length=0.0,  # Not used
                matches_played=standing.total_matches,
                victories=standing.total_victories,
                ties=standing.total_ties,
                losses=standing.total_losses,
                updated_at=standing.updated_at,
                user_name=standing.user.profile.full_name if standing.user and standing.user.profile else None,
                user_avatar=standing.user.effective_avatar_url if standing.user else None,
                is_disqualified=is_dq,  # Story 12.6
            )

            if is_dq:
                dq_items.append(item)
            else:
                non_dq_items.append(item)

        # Story 12.6: DQ users appear at bottom of leaderboard
        items = non_dq_items + dq_items

    current_phase = TATournamentPhaseAPI(
        settings.additional_rules.get("current_phase", "qualifier")
    )

    # Determine available phases based on bracket setting
    if settings.has_knockout_stage:
        available_phases = ["qualifier", "semifinal", "final"]
        if settings.has_requalification:
            available_phases.insert(1, "requalification")
    else:
        available_phases = ["qualifier"]

    # Check if knockout bracket is completed - apply final standings from knockout
    if settings.has_knockout_stage and not phase:
        from app.models.trout_area import TAKnockoutBracket
        bracket_query = select(TAKnockoutBracket).where(
            TAKnockoutBracket.event_id == event_id,
            TAKnockoutBracket.is_completed == True,
        )
        bracket_result = await db.execute(bracket_query)
        bracket = bracket_result.scalar_one_or_none()

        if bracket and bracket.final_standings:
            # Reorder items based on knockout final_standings (positions 1-4)
            final_standings = bracket.final_standings  # {"1": user_id, "2": user_id, ...}
            user_id_to_item = {item.user_id: item for item in items}

            reordered_items = []
            used_user_ids = set()

            # First, add knockout placements (1st through 4th)
            for position in ["1", "2", "3", "4"]:
                if position in final_standings:
                    user_id = final_standings[position]
                    if user_id in user_id_to_item:
                        item = user_id_to_item[user_id]
                        item.rank = int(position)  # Override rank with knockout placement
                        reordered_items.append(item)
                        used_user_ids.add(user_id)

            # Then add remaining participants (5th onwards) in original order
            remaining_rank = len(reordered_items) + 1
            for item in items:
                if item.user_id not in used_user_ids:
                    if item.rank is not None:  # Not DQ
                        item.rank = remaining_rank
                        remaining_rank += 1
                    reordered_items.append(item)

            items = reordered_items

    return {
        "items": items,
        "total": len(items),
        "phase": phase.value if phase else current_phase.value,
        "qualified_count": min(len(items), settings.knockout_qualifiers),
        "requalification_count": settings.requalification_slots if settings.has_requalification else 0,
        "has_knockout_bracket": settings.has_knockout_stage,
        "available_phases": available_phases,
    }


# =============================================================================
# Detailed Rankings Endpoints (Leg-by-Leg with Tiebreakers)
# =============================================================================

@router.get("/events/{event_id}/rankings/leg/{leg_number}")
async def get_leg_ranking(
    event_id: int,
    leg_number: int,
    request: Request,
    phase: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get cumulative ranking up to specified leg with detailed stats.

    Returns rankings with V/T/T0/L/L0 breakdown and proper tiebreaker sorting.
    """
    from app.services.ta_ranking import TARankingService
    from app.schemas.trout_area import TACompetitorDetailedStats, TALegRankingResponse

    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    ranking_service = TARankingService(db)
    rankings = await ranking_service.compute_leg_ranking(event_id, leg_number, phase)

    current_phase = TATournamentPhaseAPI(
        settings.additional_rules.get("current_phase", "qualifier") if settings else "qualifier"
    )

    return {
        "event_id": event_id,
        "leg_number": leg_number,
        "phase": current_phase,
        "is_cumulative": True,
        "rankings": rankings,
        "total_participants": len(rankings),
    }


@router.get("/events/{event_id}/rankings")
async def get_overall_ranking(
    event_id: int,
    request: Request,
    phase: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get overall ranking for the event (all legs cumulative).

    Returns rankings with V/T/T0/L/L0 breakdown and proper tiebreaker sorting.
    """
    from app.services.ta_ranking import TARankingService

    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    ranking_service = TARankingService(db)
    rankings = await ranking_service.compute_leg_ranking(event_id, phase=phase)

    current_phase = TATournamentPhaseAPI(
        settings.additional_rules.get("current_phase", "qualifier") if settings else "qualifier"
    )

    return {
        "event_id": event_id,
        "leg_number": None,
        "phase": current_phase,
        "is_cumulative": True,
        "rankings": rankings,
        "total_participants": len(rankings),
    }


@router.get("/events/{event_id}/leg-status/{leg_number}")
async def get_leg_completion_status(
    event_id: int,
    leg_number: int,
    request: Request,
    phase: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get leg completion status for standings recalculation (Story 12.1).

    Returns completion percentage and whether the leg is complete.
    Used by frontend to show leg progress and trigger standings refresh.
    """
    from app.services.ta_ranking import TARankingService

    # Verify event exists and user has access
    await get_ta_event(event_id, db, request)

    ranking_service = TARankingService(db)
    status = await ranking_service.get_leg_completion_status(event_id, leg_number, phase)

    return {
        "event_id": event_id,
        **status,
    }


@router.get("/events/{event_id}/matches/leg/{leg_number}")
async def get_leg_matches(
    event_id: int,
    leg_number: int,
    request: Request,
    phase: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get all matches for a specific leg with A vs B breakdown.

    Returns match details showing catches, outcomes, and points for both competitors.
    """
    from app.services.ta_ranking import TARankingService

    event = await get_ta_event(event_id, db, request)
    settings = event.ta_settings

    ranking_service = TARankingService(db)
    matches = await ranking_service.get_leg_matches(event_id, leg_number, phase)

    # Return the requested phase filter, or default to qualifier if not specified
    response_phase = TATournamentPhaseAPI(phase) if phase else TATournamentPhaseAPI.QUALIFIER

    return {
        "event_id": event_id,
        "leg_number": leg_number,
        "phase": response_phase,
        "matches": matches,
        "total_matches": len(matches),
    }


@router.get("/events/{event_id}/statistics")
async def get_event_statistics(
    event_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get event-level statistics for TA competition.

    Returns total participants, matches, catches, and top performers.
    """
    from app.services.ta_ranking import TARankingService

    event = await get_ta_event(event_id, db, request)

    ranking_service = TARankingService(db)
    stats = await ranking_service.get_event_statistics(event_id)

    return stats


@router.get("/events/{event_id}/standings/export")
async def export_standings_csv(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
):
    """
    Export TA standings to CSV.

    Includes:
    - QUALIFIER PHASE: Rankings from qualifier legs
    - FINAL RANKING: Overall standings with points, catches, wins/ties/losses
    """
    from fastapi.responses import StreamingResponse
    from io import StringIO
    import csv
    from app.services.ta_ranking import TARankingService

    event = await get_ta_event(event_id, db, request, require_settings=False)
    event_name_safe = sanitize_filename(event.name) if event.name else f"Event_{event_id}"
    start_date_str = event.start_date.strftime("%d%m%Y") if event.start_date else "nodate"

    # Get rankings
    ranking_service = TARankingService(db)
    rankings = await ranking_service.compute_leg_ranking(event_id)

    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)

    # === SECTION: Event Info ===
    writer.writerow(["TA Competition Results"])
    writer.writerow(["Event:", event.name or f"Event {event_id}"])
    writer.writerow(["Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    writer.writerow([])

    # === SECTION: Qualifier Phase Rankings ===
    writer.writerow(["=== QUALIFIER PHASE RANKINGS ==="])
    writer.writerow([])

    # Header
    writer.writerow([
        "Rank", "Draw #", "Name",
        "Points", "Catches", "Victories", "Ties", "Losses",
        "Matches Played", "Win Rate %"
    ])

    # Data rows
    for idx, ranking in enumerate(rankings, 1):
        matches_played = ranking.get("victories", 0) + ranking.get("ties", 0) + ranking.get("losses", 0)
        win_rate = (ranking.get("victories", 0) / matches_played * 100) if matches_played > 0 else 0

        writer.writerow([
            idx,
            ranking.get("draw_number", ""),
            ranking.get("full_name", ranking.get("user_name", f"User {ranking.get('user_id', '')}")),
            ranking.get("total_points", 0),
            ranking.get("captures", ranking.get("total_catches", 0)),
            ranking.get("victories", 0),
            ranking.get("ties", 0),
            ranking.get("losses", 0),
            matches_played,
            f"{win_rate:.1f}",
        ])

    writer.writerow([])

    # === SECTION: Final Ranking (same as qualifier for now) ===
    writer.writerow(["=== FINAL RANKING ==="])
    writer.writerow([])
    writer.writerow([
        "Position", "Name", "Total Points", "Total Catches"
    ])

    for idx, ranking in enumerate(rankings, 1):
        writer.writerow([
            idx,
            ranking.get("full_name", ranking.get("user_name", f"User {ranking.get('user_id', '')}")),
            ranking.get("total_points", 0),
            ranking.get("captures", ranking.get("total_catches", 0)),
        ])

    # Return CSV file
    output.seek(0)
    filename = f"{event_name_safe}_{start_date_str}_TA_Standings.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/events/{event_id}/team-standings")
async def get_team_standings(
    event_id: int,
    request: Request,
    phase: Optional[str] = Query(None, description="Filter by phase (qualifier, semifinal, final)"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Get team standings for TA team events.

    Returns team rankings with member breakdown.
    Only available for events with is_team_event=True.

    Scoring methods:
    - sum: Total of all member points
    - average: Average of member points
    - best_n: Sum of top N member scores

    Tiebreakers: points → captures → victories → ties_with_fish
    """
    from app.services.ta_ranking import TARankingService

    event = await get_ta_event(event_id, db, request)

    # Check if team event
    settings_query = select(TAEventSettings).where(TAEventSettings.event_id == event_id)
    settings_result = await db.execute(settings_query)
    settings = settings_result.scalar_one_or_none()

    if not settings or not settings.is_team_event:
        raise HTTPException(
            status_code=400,
            detail="Team standings only available for team events"
        )

    ranking_service = TARankingService(db)
    team_rankings = await ranking_service.compute_team_ranking(event_id, phase=phase)

    return {
        "items": team_rankings,
        "total": len(team_rankings),
        "scoring_method": settings.team_scoring_method or "sum",
        "team_size": settings.team_size,
    }


# =============================================================================
# Duration Estimate Endpoint
# =============================================================================

@router.get("/events/{event_id}/lineups/export")
async def export_lineups_excel(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
):
    """
    Export TA lineups and game cards to Excel.

    Sheet 1: Lineup - Grid (rows=legs, cols=seats) with draw numbers
    Sheet 2: User Seat Rotation - Draw#, Username, then seat for each leg
    Sheet 3: Seat-Leg Pivot - Grid (rows=seats, cols=legs) with draw numbers
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from io import BytesIO

    event = await get_ta_event(event_id, db, request, require_settings=False)

    # Get lineups
    lineup_query = (
        select(TALineup)
        .options(selectinload(TALineup.user).selectinload(UserAccount.profile))
        .where(TALineup.event_id == event_id)
        .order_by(TALineup.leg_number, TALineup.seat_number)
    )
    result = await db.execute(lineup_query)
    lineups = result.scalars().all()

    if not lineups:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No lineup data found for this event",
        )

    # Build lookup structures
    distinct_legs = sorted(set(l.leg_number for l in lineups))
    distinct_seats = sorted(set(l.seat_number for l in lineups if l.seat_number))

    # seat_draw[leg][seat] = draw_number
    # ghost_map[(leg, seat)] = is_ghost
    seat_draw: dict[int, dict[int, int]] = {}
    ghost_map: dict[tuple[int, int], bool] = {}

    # data_cards[draw_num] = { 'user_name': ..., 'legs': { leg: seat, ... } }
    data_cards: dict[int, dict] = {}

    for lineup in lineups:
        leg = lineup.leg_number
        seat = lineup.seat_number
        draw_num = lineup.draw_number

        # Build seat_draw grid
        if leg not in seat_draw:
            seat_draw[leg] = {}
        seat_draw[leg][seat] = draw_num

        # Track ghosts
        ghost_map[(leg, seat)] = lineup.is_ghost

        # Build user cards (skip ghosts)
        if not lineup.is_ghost and draw_num:
            if draw_num not in data_cards:
                user_name = (
                    f"{lineup.user.profile.last_name} {lineup.user.profile.first_name}".strip()
                    if lineup.user and lineup.user.profile else f"User {lineup.user_id}"
                )
                data_cards[draw_num] = {"user_name": user_name, "legs": {}}
            data_cards[draw_num]["legs"][leg] = seat

    # Create Excel workbook
    wb = Workbook()
    ghost_fill = PatternFill(fill_type='solid', start_color='90EE90', end_color='90EE90')
    bold_font = Font(bold=True)

    # ===================== SHEET #1: Lineup Grid (Legs x Seats) =====================
    ws_lineup = wb.active
    ws_lineup.title = "Lineup"

    # Header row: "Leg", then seats, then "Sum"
    ws_lineup.cell(row=1, column=1, value="Leg").font = bold_font
    for col_idx, seat_num in enumerate(distinct_seats, start=2):
        ws_lineup.cell(row=1, column=col_idx, value=f"#S{seat_num:02d}").font = bold_font
    sum_col = len(distinct_seats) + 2
    ws_lineup.cell(row=1, column=sum_col, value="Sum").font = bold_font

    # Data rows: one per leg, with row sum
    col_sums = {seat: 0 for seat in distinct_seats}
    for row_idx, leg_num in enumerate(distinct_legs, start=2):
        ws_lineup.cell(row=row_idx, column=1, value=leg_num)
        row_sum = 0
        for col_idx, seat_num in enumerate(distinct_seats, start=2):
            draw_val = seat_draw.get(leg_num, {}).get(seat_num, "")
            cell = ws_lineup.cell(row=row_idx, column=col_idx, value=draw_val)
            if ghost_map.get((leg_num, seat_num), False):
                cell.fill = ghost_fill
            # Add to sums (only if numeric)
            if isinstance(draw_val, (int, float)):
                row_sum += draw_val
                col_sums[seat_num] += draw_val
        # Row sum cell
        sum_cell = ws_lineup.cell(row=row_idx, column=sum_col, value=row_sum)
        sum_cell.font = bold_font

    # Column sums row at the bottom
    sum_row = len(distinct_legs) + 2
    ws_lineup.cell(row=sum_row, column=1, value="Sum").font = bold_font
    for col_idx, seat_num in enumerate(distinct_seats, start=2):
        sum_cell = ws_lineup.cell(row=sum_row, column=col_idx, value=col_sums[seat_num])
        sum_cell.font = bold_font

    # ===================== SHEET #2: User Seat Rotation =====================
    ws_rotation = wb.create_sheet(title="User Seat Rotation")

    # Header: Draw#, Username, then Leg X Seat columns
    ws_rotation.cell(row=1, column=1, value="Draw #").font = bold_font
    ws_rotation.cell(row=1, column=2, value="Username").font = bold_font
    for col_idx, leg_num in enumerate(distinct_legs, start=3):
        ws_rotation.cell(row=1, column=col_idx, value=f"Leg {leg_num} Seat").font = bold_font

    # Data rows: one per user (sorted by draw number)
    sorted_draw_numbers = sorted(data_cards.keys())
    for row_idx, draw_num in enumerate(sorted_draw_numbers, start=2):
        info = data_cards[draw_num]
        ws_rotation.cell(row=row_idx, column=1, value=draw_num)
        ws_rotation.cell(row=row_idx, column=2, value=info["user_name"])

        for col_idx, leg_num in enumerate(distinct_legs, start=3):
            seat_val = info["legs"].get(leg_num, "")
            ws_rotation.cell(row=row_idx, column=col_idx, value=seat_val)

    # ===================== SHEET #3: Seat-Leg Pivot (Seats x Legs) =====================
    ws_pivot = wb.create_sheet(title="Seat-Leg Pivot")

    # Header: "Seat", then Leg columns, then "Sum"
    ws_pivot.cell(row=1, column=1, value="Seat").font = bold_font
    for col_idx, leg_num in enumerate(distinct_legs, start=2):
        ws_pivot.cell(row=1, column=col_idx, value=f"Leg {leg_num}").font = bold_font
    pivot_sum_col = len(distinct_legs) + 2
    ws_pivot.cell(row=1, column=pivot_sum_col, value="Sum").font = bold_font

    # Data rows: one per seat, with row sum
    pivot_col_sums = {leg: 0 for leg in distinct_legs}
    for row_idx, seat_num in enumerate(distinct_seats, start=2):
        ws_pivot.cell(row=row_idx, column=1, value=f"Seat {seat_num}")
        row_sum = 0
        for col_idx, leg_num in enumerate(distinct_legs, start=2):
            draw_val = seat_draw.get(leg_num, {}).get(seat_num, "")
            cell = ws_pivot.cell(row=row_idx, column=col_idx, value=draw_val)
            if ghost_map.get((leg_num, seat_num), False):
                cell.fill = ghost_fill
            # Add to sums (only if numeric)
            if isinstance(draw_val, (int, float)):
                row_sum += draw_val
                pivot_col_sums[leg_num] += draw_val
        # Row sum cell
        sum_cell = ws_pivot.cell(row=row_idx, column=pivot_sum_col, value=row_sum)
        sum_cell.font = bold_font

    # Column sums row at the bottom
    pivot_sum_row = len(distinct_seats) + 2
    ws_pivot.cell(row=pivot_sum_row, column=1, value="Sum").font = bold_font
    for col_idx, leg_num in enumerate(distinct_legs, start=2):
        sum_cell = ws_pivot.cell(row=pivot_sum_row, column=col_idx, value=pivot_col_sums[leg_num])
        sum_cell.font = bold_font

    # Save to BytesIO and return
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    event_name_safe = sanitize_filename(event.name) if event.name else f"Event_{event_id}"
    start_date_str = event.start_date.strftime("%d%m%Y") if event.start_date else "nodate"
    filename = f"{event_name_safe}_{start_date_str}_TA_Lineup.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/events/{event_id}/game-cards/export")
async def export_game_cards_csv(
    event_id: int,
    request: Request,
    current_user: UserAccount = Depends(EventOwnerOrAdmin()),
    db: AsyncSession = Depends(get_db),
):
    """
    Export TA game cards to CSV.

    Includes all game cards with owner details, catches, validation status,
    and dispute information.
    """
    from fastapi.responses import StreamingResponse
    from io import StringIO
    import csv

    event = await get_ta_event(event_id, db, request, require_settings=False)
    event_name_safe = sanitize_filename(event.name) if event.name else f"Event_{event_id}"
    start_date_str = event.start_date.strftime("%d%m%Y") if event.start_date else "nodate"

    # Get all game cards with user details and match info
    cards_query = (
        select(TAGameCard)
        .options(
            selectinload(TAGameCard.user).selectinload(UserAccount.profile),
            selectinload(TAGameCard.opponent).selectinload(UserAccount.profile),
            selectinload(TAGameCard.match),
        )
        .where(TAGameCard.event_id == event_id)
        .order_by(TAGameCard.leg_number, TAGameCard.my_seat)
    )
    result = await db.execute(cards_query)
    cards = result.scalars().all()

    # Create CSV content
    output = StringIO()
    writer = csv.writer(output)

    # === SECTION: Event Info ===
    writer.writerow(["TA Game Cards Export"])
    writer.writerow(["Event:", event.name or f"Event {event_id}"])
    writer.writerow(["Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M")])
    writer.writerow([])

    # === SECTION: Game Cards ===
    writer.writerow(["=== GAME CARDS ==="])
    writer.writerow([])

    # Simplified header - most useful columns for organizers
    writer.writerow([
        "Leg",
        "Phase",
        "Player Name",
        "Catches",
        "Opponent Name",
        "Opponent Catches",
        "Result",
        "Status",
        "Disputed",
        "Notes"
    ])

    # Data rows
    for card in cards:
        # Get owner name
        owner_name = ""
        if card.user and card.user.profile:
            owner_name = f"{card.user.profile.last_name} {card.user.profile.first_name}".strip()
        elif card.user:
            owner_name = f"User {card.user_id}"

        # Get opponent name
        opponent_name = ""
        if card.is_ghost_opponent:
            opponent_name = "[BYE]"
        elif card.opponent and card.opponent.profile:
            opponent_name = f"{card.opponent.profile.last_name} {card.opponent.profile.first_name}".strip()
        elif card.opponent:
            opponent_name = f"User {card.opponent_id}"

        # Get phase from match
        phase = "Qualifier"
        if card.match and card.match.phase:
            phase_map = {
                "qualifier": "Qualifier",
                "requalification": "Requalification",
                "semifinal": "Semifinal",
                "final_grand": "Grand Final",
                "final_small": "Small Final (3rd/4th)",
            }
            phase = phase_map.get(card.match.phase, card.match.phase)

        # Determine result
        result = ""
        my_catches = card.my_catches or 0
        opp_catches = card.opponent_catches or 0
        if card.is_validated:
            if card.is_ghost_opponent:
                result = "WIN (Bye)"
            elif my_catches > opp_catches:
                result = "WIN"
            elif my_catches < opp_catches:
                result = "LOSS"
            else:
                result = "TIE"
        elif card.is_submitted:
            result = "Pending validation"
        else:
            result = "Not submitted"

        # Notes (dispute reason if any)
        notes = card.dispute_reason or ""

        writer.writerow([
            card.leg_number,
            phase,
            owner_name,
            card.my_catches if card.my_catches is not None else "",
            opponent_name,
            card.opponent_catches if card.opponent_catches is not None else "",
            result,
            card.status or "",
            "Yes" if card.is_disputed else "",
            notes,
        ])

    writer.writerow([])

    # === SECTION: Summary Statistics ===
    writer.writerow(["=== SUMMARY ==="])
    writer.writerow([])

    total_cards = len(cards)
    submitted_cards = sum(1 for c in cards if c.is_submitted)
    validated_cards = sum(1 for c in cards if c.is_validated)
    disputed_cards = sum(1 for c in cards if c.is_disputed)
    ghost_cards = sum(1 for c in cards if c.is_ghost_opponent)

    writer.writerow(["Total Game Cards:", total_cards])
    writer.writerow(["Submitted Cards:", submitted_cards])
    writer.writerow(["Validated Cards:", validated_cards])
    writer.writerow(["Disputed Cards:", disputed_cards])
    writer.writerow(["Ghost Opponent Cards:", ghost_cards])

    # Return CSV file
    output.seek(0)
    filename = f"{event_name_safe}_{start_date_str}_TA_GameCards.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/duration-estimate", response_model=TADurationEstimateResponse)
async def estimate_duration(
    data: TADurationEstimateRequest,
) -> dict:
    """
    Calculate estimated duration for a TA event.

    This is a utility endpoint that doesn't require authentication.
    Useful for planning events before creating them.
    """
    algorithm = map_pairing_algorithm(data.algorithm)

    result = TAPairingService.calculate_event_duration(
        num_participants=data.num_participants,
        algorithm=algorithm,
        match_duration_minutes=data.match_duration_minutes,
        break_between_rounds_minutes=data.break_between_rounds_minutes,
        custom_rounds=data.custom_rounds,
    )

    return result
